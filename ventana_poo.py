import customtkinter as ctk
import serial
import csv
import time
from tkinter import ttk
import tkinter as tk
from CTkMenuBarPlus import *
from datetime import datetime
from CTkMessagebox import CTkMessagebox
from tkinter.filedialog import asksaveasfilename, askopenfilename
from PIL import Image, ImageTk
import serial.tools.list_ports
import os
from customtkinter import CTkOptionMenu
import openpyxl
import threading
import re
from openpyxl.styles import PatternFill, Font, Alignment
from tkinter import filedialog
import pandas as pd
import matplotlib.pyplot as plt
import io

global Lectura_barometro, LecturaT1, LecturaT2, LecturaH1, LecturaH2
Lectura_barometro = "0.0"
LecturaT1 = "0.0"
LecturaT2 = "0.0"
LecturaH1 = "0.0"
LecturaH2 = "0.0"
barometro_ser = None
higrometro_ser = None


def conectar_higrometro(puerto):
    global higrometro_ser
    try:
        if higrometro_ser and higrometro_ser.is_open:
            higrometro_ser.close()
        time.sleep(0.5)
        higrometro_ser = serial.Serial(puerto, 9600, timeout=1)
        higrometro_ser.reset_input_buffer()
        print(f"Puerto {puerto} abierto con éxito")
        return True
    except Exception as e:
        print(f"Error al abrir puerto higrómetro: {e}")
        return False


def conectar_barometro(puerto):
    global barometro_ser
    try:
        if barometro_ser is not None:
            try:
                barometro_ser.close()
                time.sleep(0.5)
            except:
                pass

        barometro_ser = serial.Serial(
            port=puerto,
            baudrate=9600,
            parity=serial.PARITY_NONE,
            stopbits=serial.STOPBITS_ONE,
            bytesize=serial.EIGHTBITS,
            timeout=1,
            rtscts=False,
            dsrdtr=False,
        )

        barometro_ser.reset_input_buffer()
        barometro_ser.reset_output_buffer()

        print(f"Conexión BAR exitosa en {puerto}")
        return True
    except serial.SerialException as e:
        print(f"Error de hardware/permisos en puerto: {e}")
        barometro_ser = None
        return False
    except Exception as e:
        print(f"Error inesperado bar - {e}")
        return False


def read_hidro():
    global LecturaT1, LecturaT2, LecturaH1, LecturaH2, higrometro_ser

    # print("Hilo higro")

    while True:
        if higrometro_ser and higrometro_ser.is_open:
            try:
                dato_raw = higrometro_ser.readline()
                if dato_raw:
                    cadena = dato_raw.decode("ascii", errors="ignore").strip()

                    numeros = re.findall(r"[-+]?\d*\.\d+|\d+", cadena)
                    if len(numeros) >= 9:
                        LecturaT1 = numeros[6]
                        LecturaH1 = numeros[7]
                        LecturaT2 = numeros[8]
                        LecturaH2 = numeros[9]
            except Exception as e:
                print(f"Error procesando cadena larga: {e}")

        time.sleep(1)  # Frecuencia de actualización


def read_bar():
    global Lectura_barometro, barometro_ser
    while True:
        if barometro_ser and barometro_ser.is_open:
            try:
                barometro_ser.reset_input_buffer()
                barometro_ser.write(b":clr\n:SENS:PRES?\n")

                time.sleep(0.2)
                lectura_1 = barometro_ser.readline()

                if lectura_1:
                    l2 = lectura_1.decode("ascii", errors="ignore").strip()
                    import re

                    nums = re.findall(r"[-+]?\d*\.\d+|\d+", l2)
                    if nums:
                        Lectura_barometro = "{:.2f}".format(float(nums[0]))

                time.sleep(0.8)
            except Exception as e:
                print(f"Error en hilo barómetro: {e}")
                time.sleep(1)
        else:
            time.sleep(1)


def ruta(r_relativa):
    import sys, os

    try:
        b_path = sys._MEIPASS
    except Exception:
        b_path = os.path.abspath(".")
    return os.path.join(b_path, r_relativa)


class Comunication:
    """
    Establecer la comunicación serial con MT
    """

    def __init__(self, port: str, baud: int):
        self.port = port
        self.baud = baud
        self.ser = None
        self.conectado = True

    def conexion(self):
        """
        metodo de conexion
        """
        if self.ser and self.ser.is_open:
            self.ser.close()

        try:
            self.ser = serial.Serial(
                port=self.port,
                baudrate=self.baud,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                timeout=0.5,
            )
            time.sleep(2)
            self.ser.reset_input_buffer()
            print(f"M-Conexion: {self.port}")
            self.conectado = True
            return True
        except Exception as e:
            self.conectado = False
            print(f"Error: M-Conexion: {e}")
            return False

    def envio(self, data: str):
        """
        Envio generico
        """
        if self.ser and self.ser.is_open:
            try:
                self.ser.write((data + "\r\n").encode("ascii"))
                # print("M-Send")
            except Exception as e:
                print(f"Error M-Send: {e}")
        # else:
        # print("P cerrado")

    def respuesta(self):
        """
        Respuesta del write ascii
        """
        if self.ser and self.ser.is_open:
            return self.ser.readline().decode("ascii", errors="ignore").strip()

    def cerrar_puerto(self):
        if self.ser and self.ser.is_open:
            self.ser.close()


class Shell:
    """
    Manda las intrucciones zero, tara, peso
    """

    def __init__(self, comando: Comunication):
        self.comando = comando

    def peso_instantaneo(self):
        """
        Envia el peso al instante este o no estable
        """
        self.comando.envio("SI")
        lecture = self.comando.respuesta()
        if lecture:
            return lecture.split()
        return []

    def zero(self):
        """
        Envia el cero
        """
        self.comando.envio("Z")
        lecture = self.comando.respuesta()
        return print("Zero enviado")

    def tara(self):
        """
        Envia la tara
        """
        self.comando.envio("T")
        self.comando.respuesta()
        return print("Tara enviada")

    def quitar_tara(self):
        """
        Quita la tara si es que tiene
        """
        self.comando.envio("TAC")
        self.comando.respuesta()
        return print("Quite Tara")

    def peso_estable(self):
        """
        Envia el peso cuando esta estable
        """
        self.comando.envio("S")
        lecture = self.comando.respuesta()
        print("peso estable")
        if "S +" in lecture:
            return "OVERLOAD"
        if lecture:
            return lecture.split()
        return []

    def cali_cero(self):
        try:
            self.comando.envio("Z")
            lecture = self.comando.respuesta()
            return lecture
        except Exception as e:
            print(f"Error cali cero {e}")

    def consulta_datos(self):
        try:
            self.comando.envio("I2")
            lecture = self.comando.respuesta()
            return lecture
        except Exception as e:
            print(f"errro{e}")

    def leer_variable(self, indice):
        try:
            comando_i = f"R{indice}\r\n"
            self.comando.ser.write(comando_i.encode("ascii"))
            respuesta = self.comando.respuesta()

            if respuesta:
                partes = respuesta.split()
                if len(partes) >= 2:
                    return True, partes[1]
            return False, None
        except Exception as e:
            print(f"Error RR {e}")
            return False, None

    def escribir_variable(self, indice, valor):
        try:
            comando_i = f"W{indice} {valor}\r\n"
            self.comando.ser.write(comando_i.encode("ascii"))
            respuesta = self.comando.ser.read(1)

            if respuesta and respuesta[0] == 0x06:
                print(f"good Variable {indice}, {valor}")
                return True
            else:
                print(f"error {indice}")
                return False
        except Exception as e:
            print(f"Error escribir_v {e}")
            return False

    def calibrar_cero(self):
        lecture = self.peso_instantaneo()
        if len(lecture) >= 3:
            peso_actual = float(lecture[2])
            estado = lecture[1]

            if abs(peso_actual) > 5:
                return False, f"No vacia"
            if estado != "S":
                return False, "No estable"

        exito = self.escribir_variable(40, 1)
        if exito:
            time.sleep(2)

            lecture = self.peso_instantaneo()
            if len(lecture) >= 3:
                peso_final = float(lecture[2])
                return True, f"Calibrado {peso_final}"
            else:
                return True, f"Calibraco goos"
        else:
            return False, "Error calibracion"

    def obtener_tara(self):
        if self.comando.ser and self.comando.ser.is_open:
            try:
                self.comando.envio("TA")
                lecture = self.comando.respuesta()
                partes = lecture.split()

                if len(partes) >= 4:
                    return f"{partes[2]} {partes[3]}"
                return "0.0"
            except Exception as e:
                return f"error {e}"

        return "--"

    def peso_estable_2(self):
        self.comando.envio("S")
        lecture = self.comando.respuesta()
        if lecture is None:
            return []
        if "S +" in lecture:
            return "OVERLOAD"
        return lecture.split()


class Repetibilidad(ctk.CTkToplevel):
    def __init__(self, master, shell: Shell, lista_instrumentos, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.lista_instrumentos = lista_instrumentos
        self.shell = shell
        self.lista_pesas_completa = []
        self.title("Repetibilidad")
        self.geometry("800x800+190+20")
        self.grab_set()

        self.menu()
        self.widgets()
        self.actualizar_peso()

    def widgets(self):
        self.frame_repetibilidad = ctk.CTkFrame(self, width=800, height=700)
        self.frame_repetibilidad.pack(fill="both", expand=True, padx=5, pady=5)
        self.frame_repetibilidad.pack_propagate(False)

        self.frame_peso = ctk.CTkFrame(self.frame_repetibilidad, width=400, height=150)
        self.frame_peso.pack(anchor="n", padx=10, pady=10, fill="x")
        self.frame_peso.pack_propagate(False)

        self.frame_peso_instantaneo = ctk.CTkFrame(
            self.frame_peso, width=450, height=150
        )
        self.frame_peso_instantaneo.pack(side="left", padx=10, pady=10)
        self.frame_peso_instantaneo.pack_propagate(False)

        self.label_unidad = ctk.CTkLabel(
            self.frame_peso_instantaneo, text="kg", font=("Cambria", 50)
        )
        self.label_unidad.pack(side="right", padx=20, pady=5)

        self.frame_peso_al_momento = ctk.CTkFrame(
            self.frame_peso_instantaneo, width=450, height=150
        )
        self.frame_peso_al_momento.pack(side="left", padx=10, pady=10)
        self.frame_peso_al_momento.pack_propagate(False)

        self.label_peso = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 70)
        )
        self.label_peso.pack(padx=20, pady=5, expand=True)

        self.label_tara = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 12)
        )
        self.label_tara.place(x=8, y=8)

        self.label_neto = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 12)
        )
        self.label_neto.place(x=8, y=79)

        # ---------BOTONES-------------------
        imagen_tarar = Image.open(ruta(os.path.join("Icon", "aumentar.png")))
        icon_tarar = ctk.CTkImage(
            light_image=imagen_tarar, dark_image=imagen_tarar, size=(30, 30)
        )

        self.button_tarar = ctk.CTkButton(
            self.frame_peso,
            text="Tara",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.tara,
            image=icon_tarar,
            compound="top",
        )
        self.button_tarar.pack(side="left", expand=True)
        # ---------------BOTON ZERO-------------
        imagen_zero = Image.open(ruta(os.path.join("Icon", "cero.png")))
        icon_zero = ctk.CTkImage(
            light_image=imagen_zero, dark_image=imagen_zero, size=(30, 30)
        )

        self.button_zero = ctk.CTkButton(
            self.frame_peso,
            text="Zero",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.zero,
            image=icon_zero,
            compound="top",
        )
        self.button_zero.pack(side="left", expand=True)

        # ---------------BOTON QUITAR TARAR-------------
        imagen_q_tara = Image.open(ruta(os.path.join("Icon", "perdida-peso.png")))
        icon_q_tara = ctk.CTkImage(
            light_image=imagen_q_tara, dark_image=imagen_q_tara, size=(30, 30)
        )
        self.button_quitar_tara = ctk.CTkButton(
            self.frame_peso,
            text="Eliminar Tara",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.quitar_tara,
            image=icon_q_tara,
            compound="top",
        )
        self.button_quitar_tara.pack(side="left", expand=True)

        # ----------------Condiciones Ambientales-----------------
        self.frame_condiciones = ctk.CTkFrame(
            self.frame_repetibilidad, width=400, height=150
        )
        self.frame_condiciones.pack(anchor="s", padx=10, pady=5, fill="x")
        self.frame_condiciones.pack_propagate(False)

        self.frame_condicion1 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion1.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion1.pack_propagate(False)

        self.label_condiciones = ctk.CTkLabel(
            self.frame_condicion1,
            text="Presión",
            font=("Cambria", 20),
        )
        self.label_condiciones.pack(anchor="n", padx=2, pady=5, expand=True)

        self.frame_condicion2 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion2.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion2.pack_propagate(False)

        self.label_condiciones2 = ctk.CTkLabel(
            self.frame_condicion2,
            text="Humedad",
            font=("Cambria", 20),
        )
        self.label_condiciones2.pack(anchor="n", padx=2, pady=5, expand=True)

        self.frame_condicion3 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion3.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion3.pack_propagate(False)

        self.label_condiciones3 = ctk.CTkLabel(
            self.frame_condicion3,
            text="Temperatura",
            font=("Cambria", 20),
        )
        self.label_condiciones3.pack(anchor="n", padx=2, pady=5, expand=True)
        # ---------PRESION-----------------

        self.frame_presion = ctk.CTkFrame(self.frame_condicion1, width=200, height=50)
        self.frame_presion.pack(expand=True, anchor="n")
        self.frame_presion.pack_propagate(False)

        self.label_presion = ctk.CTkLabel(
            self.frame_presion,
            text="---",
            font=("Cambria", 30),
        )
        self.label_presion.pack(expand=True)

        self.combo_presion = ctk.CTkOptionMenu(
            master=self.frame_condicion1,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_presion.pack(expand=True)
        self.combo_presion.set("Modelos")

        # -------------HIGROMETRO-----------------

        self.frame_higrometro = ctk.CTkFrame(
            self.frame_condicion2, width=200, height=50
        )
        self.frame_higrometro.pack(expand=True, anchor="n")
        self.frame_higrometro.pack_propagate(False)

        self.label_higrometro = ctk.CTkLabel(
            self.frame_higrometro,
            text="---",
            font=("Cambria", 30),
        )
        self.label_higrometro.pack(expand=True)
        self.combo_higrometro = ctk.CTkOptionMenu(
            master=self.frame_condicion2,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_higrometro.pack(expand=True)
        self.combo_higrometro.set("Modelos")

        # --------------Temperatura-----------------

        self.frame_temperatura = ctk.CTkFrame(
            self.frame_condicion3, width=200, height=50
        )
        self.frame_temperatura.pack(expand=True, anchor="n")
        self.frame_temperatura.pack_propagate(False)

        self.label_temperatura = ctk.CTkLabel(
            self.frame_temperatura, text="---", font=("Cambria", 30)
        )
        self.label_temperatura.pack(expand=True)
        self.combo_temperatura = ctk.CTkOptionMenu(
            master=self.frame_condicion3,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_temperatura.pack(expand=True)
        self.combo_temperatura.set("Modelos")

        # --------------------------------
        self.frame_registro = ctk.CTkFrame(
            self.frame_repetibilidad, width=400, height=420
        )
        self.frame_registro.pack(anchor="s", padx=10, pady=5, fill="x")
        self.frame_registro.pack_propagate(False)

        self.style_tabla = ttk.Style()
        self.style_tabla.theme_use("clam")
        self.style_tabla.configure(
            "Treeview",
            background="#BD9D9D",
            foreground="black",
            rowheight=25,
            fieldbackground="#DFF2F5",
            font=("Cambria", 12),
        )
        self.style_tabla.map(
            "Treeview",
            background=[("selected", "#158A30")],
            foreground=[("selected", "white")],
        )

        self.tabla = ttk.Treeview(
            self.frame_registro,
            columns=("#", "Medicion"),
            show="headings",
            height=8,
        )
        self.tabla.heading("#", text="#")
        self.tabla.heading("Medicion", text="Indicación")

        self.tabla.column("#", width=30, anchor="center")
        self.tabla.column("Medicion", width=100, anchor="center")

        self.tabla.pack(side="left", padx=20, pady=20, anchor="n")

        # -------------Desviacion estandar----------------

        self.label_desviacion = ctk.CTkLabel(
            self.frame_registro, text="", font=("Cambria", 12)
        )
        self.label_desviacion.place(x=50, y=250)

        # ---------tabla de pesas----------------

        # ----------------BOTONOES----------
        # ----------------------------------
        self.frame_botones_pesas = ctk.CTkFrame(
            self.frame_registro, width=200, height=400
        )
        self.frame_botones_pesas.pack(side="right", padx=10, pady=10, anchor="n")
        self.frame_botones_pesas.pack_propagate(False)

        imagen_btn_check = Image.open(ruta(os.path.join("Icon", "escoger.png")))
        icon_btn_check = ctk.CTkImage(
            light_image=imagen_btn_check, dark_image=imagen_btn_check, size=(30, 30)
        )
        self.btn_check = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Seleccionar Pesas",
            command=self.mostrar_selector_pesas,
            width=200,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_btn_check,
            # compound="top",
        )
        self.btn_check.pack(side="top", expand=True, padx=10, pady=10)
        self.btn_check.pack_propagate(False)
        # --------------------------------
        imagen_eliminar_ultimo = Image.open(ruta(os.path.join("Icon", "borrar.png")))
        # imagen_eliminar_ultimo = Image.open(r"Icon/borrar.png")
        icono_eliminar_ultimo = ctk.CTkImage(
            light_image=imagen_eliminar_ultimo,
            dark_image=imagen_eliminar_ultimo,
            size=(30, 30),
        )

        self.btn_limpiar = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Limpiar Selección",
            width=200,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.eliminar_ultimo,
            image=icono_eliminar_ultimo,
            # compound="top",
        )
        self.btn_limpiar.pack(side="top", expand=True, padx=10, pady=10)

        self.tabla_sel = ttk.Treeview(
            self.frame_registro,
            columns=("#", "ID", "serie", "valor", "unidad"),
            show="headings",
            height=8,
        )

        imagen_registro = Image.open(ruta(os.path.join("Icon", "boton-agregar.png")))
        # imagen_registro = Image.open(r"Icon\boton-agregar.png")
        icon_registro = ctk.CTkImage(
            light_image=imagen_registro, dark_image=imagen_registro, size=(30, 30)
        )
        self.btn_peso = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Registrar Peso",
            width=200,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.registrar_medicion,
            image=icon_registro,
            # compound="top",
        )
        self.btn_peso.pack(side="top", expand=True, padx=10, pady=10)

        imagen_guardar = Image.open(
            ruta(os.path.join("Icon", "guardar-el-archivo.png"))
        )
        icon_save = ctk.CTkImage(
            light_image=imagen_guardar, dark_image=imagen_guardar, size=(30, 30)
        )
        self.btn_guardar = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Guardar",
            width=200,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_save,
            command=self.guardar_xlsx,
        )
        self.btn_guardar.pack(side="top", expand=True, padx=10, pady=10)
        self.btn_guardar.pack_propagate(False)

        imagen_eliminar_peso = Image.open(ruta(os.path.join("Icon", "eliminar.png")))
        icon_eliminar_peso = ctk.CTkImage(
            light_image=imagen_eliminar_peso,
            dark_image=imagen_eliminar_peso,
            size=(30, 30),
        )

        self.btn_eliminar_medicion = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Eliminar indicación",
            width=200,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.eliminar_peso,
            image=icon_eliminar_peso,
        )
        self.btn_eliminar_medicion.pack(side="top", expand=True, padx=10, pady=10)

        imagen_C = Image.open(ruta(os.path.join("Icon", "escribir.png")))
        icon_C = ctk.CTkImage(light_image=imagen_C, dark_image=imagen_C, size=(30, 30))

        self.btn_CI = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Registrar CI",
            width=200,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            command=self.registrar_CI,
            image=icon_C,
        )

        self.btn_CI.pack(side="top", expand=True, padx=10, pady=10)

        self.btn_CF = ctk.CTkButton(
            self.frame_botones_pesas,
            text="Registrar CF",
            width=200,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            command=self.registrar_CF,
            image=icon_C,
        )

        self.btn_CF.pack(side="top", expand=True, padx=10, pady=10)
        # -------------------------------------------

        self.tabla_sel.heading("#", text="#")
        self.tabla_sel.heading("ID", text="Identificación")
        self.tabla_sel.heading("serie", text="Serie")
        self.tabla_sel.heading("valor", text="V nominal")
        self.tabla_sel.heading("unidad", text="Unidad")

        self.tabla_sel.column("#", width=30, anchor="center")
        self.tabla_sel.column("ID", width=90, anchor="center")
        self.tabla_sel.column("serie", width=90, anchor="center")
        self.tabla_sel.column("valor", width=90, anchor="center")
        self.tabla_sel.column("unidad", width=70, anchor="center")
        self.tabla_sel.pack(side="right", padx=5, pady=20, anchor="n")

        self.label_suma_nominal = ctk.CTkLabel(
            self.frame_registro, text="Suma Nominal: 0 kg", font=("Cambria", 12)
        )
        self.label_suma_nominal.place(x=430, y=250)

        # _------------------condiciones----
        self.frame_R_condiciones = ctk.CTkFrame(
            self.frame_registro, width=450, height=45
        )
        self.frame_R_condiciones.place(x=50, y=290)
        self.frame_R_condiciones.pack_propagate(False)

        self.frame_r_bar = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_bar.pack(side="left", expand=True, padx=10)
        self.frame_r_bar.pack_propagate(False)

        self.label_barI = ctk.CTkLabel(
            self.frame_r_bar, text="--", font=("Cambria", 15)
        )
        self.label_barI.pack(expand=True)

        self.frame_r_pres = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_pres.pack(side="left", expand=True)
        self.frame_r_pres.pack_propagate(False)

        self.label_higI = ctk.CTkLabel(
            self.frame_r_pres, text="--", font=("Cambria", 15)
        )
        self.label_higI.pack(expand=True)

        self.frame_r_tem = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_tem.pack(side="left", expand=True)
        self.frame_r_tem.pack_propagate(False)

        self.label_temI = ctk.CTkLabel(
            self.frame_r_tem, text="--", font=("Cambria", 15)
        )
        self.label_temI.pack(expand=True)

        # ---------condiciones finales ------

        self.frame_cf = ctk.CTkFrame(self.frame_registro, width=450, height=45)
        self.frame_cf.place(x=50, y=350)
        self.frame_cf.pack_propagate(False)

        self.frame_barF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_barF.pack(side="left", expand=True, padx=10)
        self.frame_barF.pack_propagate(False)

        self.label_barF = ctk.CTkLabel(
            self.frame_barF, text="---", font=("Cambria", 15)
        )
        self.label_barF.pack(expand=True)

        self.frame_higF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_higF.pack(side="left", expand=True)
        self.frame_higF.pack_propagate(False)
        self.label_higF = ctk.CTkLabel(
            self.frame_higF, text="---", font=("Cambria", 15)
        )
        self.label_higF.pack(expand=True)

        self.frame_temF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_temF.pack(side="left", expand=True)
        self.frame_temF.pack_propagate(False)

        self.label_temF = ctk.CTkLabel(
            self.frame_temF, text="---", font=("Cambria", 15)
        )
        self.label_temF.pack(expand=True)

    def menu(self):
        self.menu = CTkMenuBar(master=self)
        self.b1 = self.menu.add_cascade("Archivo", command=self.abrir_modelos)
        self.b2_puerto_bar = self.menu.add_cascade("Barometro")
        self.dropdown_puerto_bar = CustomDropdownMenu(widget=self.b2_puerto_bar)
        self.b3_puerto_higro = self.menu.add_cascade("Higrometro")
        self.dropdown_puerto_higro = CustomDropdownMenu(widget=self.b3_puerto_higro)

        self.leer = self.menu.add_cascade("Leer xlsx", command=self.abrir_xlsx)

        self.prueba = self.menu.add_cascade("Prueba", command=self.nueva_prueba)
        self.sustitucion = self.menu.add_cascade(
            "Sustitución", command=self.mostrar_selector_sustitucion
        )

        puertos = serial.tools.list_ports.comports()
        if not puertos:
            self.dropdown_puerto_bar.add_command(label="No se encontraron puertos")
            self.dropdown_puerto_higro.add_command(label="No se encontraron puertos")
        else:
            for i in puertos:
                self.dropdown_puerto_bar.add_option(
                    option=i.device,
                    command=lambda puerto=i.device: self.seleccionar_puerto_bar(puerto),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                )
                self.dropdown_puerto_higro.add_option(
                    option=i.device,
                    command=lambda puerto=i.device: self.seleccionar_puerto_higro(
                        puerto
                    ),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                )

    def actualizar_peso(self):
        global Lectura_barometro, LecturaT1, LecturaT2
        try:
            val_menos = "{:.2f}".format(float(Lectura_barometro))
            self.label_presion.configure(text=f"{val_menos} Pa")
            self.label_temperatura.configure(text=f"{LecturaT1} °C")
            self.label_higrometro.configure(text=f"{LecturaH1} %")
        except Exception as e:
            print(f"error -{e}")
            # self.label_presion.configure(text="0.00 Pa")
            # self.label_temperatura.configure(text="0.00 °C")
            # self.label_higrometro.configure(text="0.00 %")

        lecture = self.shell.peso_instantaneo()

        if len(lecture) >= 3:
            try:
                peso_r = int(float(lecture[2]))
                self.label_peso.configure(text=str(peso_r))
                self.label_neto.configure(text=f"Neto: {peso_r} kg", text_color="white")
            except ValueError:
                pass

        # if "+" in lecture:
        #     self.label_peso.configure(text="----", text_color="#B52C19")

        #     if not self.alerta_peso_maximo:
        #         self.alerta_peso_maximo = True
        #         CTkMessagebox(title="Maximo peso", message="Sobrecarga", icon="warning")
        #     elif len(lecture) >= 3:
        #         self.alerta_peso_maximo = False

        if len(lecture) >= 3:
            wg = lecture[2]
            self.label_peso.configure(text=wg)

            if lecture[1] == "S":
                self.label_peso.configure(text_color="#158A30")
            else:
                self.label_peso.configure(text_color="#B52C19")

            valor_tara = self.shell.obtener_tara()
            self.label_tara.configure(text=f"Tara: {valor_tara}", text_color="#FF6021")

        self.after(500, self.actualizar_peso)

    def abrir_modelos(self):
        ruta = askopenfilename(
            title="Seleccionar archivo", filetypes=[("Archivos de Excel", "*.xlsx")]
        )

        if ruta:
            try:
                wb = openpyxl.load_workbook(ruta, data_only=True)

                if "Equipo CA" in wb.sheetnames:
                    hoja = wb["Equipo CA"]
                    self.equipos_ca = {}
                else:
                    CTkMessagebox(
                        title="Error",
                        message='La hoja "Equipo CA" no se encuentra en el archivo',
                        icon="error",
                    )
                    return

                lista_sensores = []
                for fila in hoja.iter_rows(min_row=2, max_col=10, values_only=True):
                    if fila[0] is not None:
                        a = str(fila[0]).strip()
                        b = str(fila[1]).strip() if fila[1] is not None else ""
                        c = str(fila[2]).strip() if fila[2] is not None else ""
                        d = str(fila[3]).strip() if fila[3] is not None else ""
                        e = str(fila[5]).strip() if fila[5] is not None else ""
                        h = str(fila[7]).strip() if fila[7] is not None else ""
                        sensor_info = f"{d}, {e}, {h}".strip(", ")

                        self.equipos_ca[sensor_info] = {
                            "key": a,
                            "magnitud": b,
                            "indicacion": c,
                            "marca": d,
                            "modelo": e,
                            "serie_sensor": h,
                        }

                        if sensor_info not in lista_sensores:
                            lista_sensores.append(sensor_info)
                self.lista_medidores = lista_sensores
                self.combo_presion.configure(values=lista_sensores)
                self.combo_higrometro.configure(values=lista_sensores)
                self.combo_temperatura.configure(values=lista_sensores)

                if "Equipo Pesas" in wb.sheetnames:
                    hoja_pesas = wb["Equipo Pesas"]
                    self.pesas = {}  # Diccionario maestro
                    self.lista_pesas_completa = []

                    # for idx, i in enumerate(
                    #     hoja_pesas.iter_rows(min_row=2, max_col=15, values_only=True)
                    # ):
                    for idx, i in enumerate(
                        hoja_pesas.iter_rows(min_row=2, max_col=15, values_only=True)
                    ):
                        if i[2] is not None:
                            # if i[2] is not None:
                            a = str(i[0]).strip()
                            b = str(i[1]).strip() if i[1] is not None else ""
                            c = str(i[2]).strip() if i[2] is not None else ""
                            e = str(i[4]).strip() if i[4] is not None else ""
                            f = str(i[5]).strip() if i[5] is not None else ""
                            g = str(i[6]).strip() if i[6] is not None else ""
                            h = str(i[7]).strip() if i[7] is not None else ""
                            j = str(i[9]).strip() if i[9] is not None else ""
                            k = str(i[10]).strip() if i[10] is not None else ""
                            L = str(i[11]).strip() if i[11] is not None else ""
                            pesa_key = f"{c} | {f} | {j}{k}"

                            self.pesas[pesa_key] = {
                                "key": a,
                                "magnitud": b,
                                "id": c,
                                "modelo": e,
                                "serie": f,
                                "juego": g,
                                "id_pesa": h,
                                "nominal": j,
                                "unidad": k,
                                "sustitucion": L,
                            }
                            self.lista_pesas_completa.append(pesa_key)

                CTkMessagebox(
                    title="Exito",
                    message="Modelos cargados correctamente.",
                    icon="check",
                )

            except Exception as e:
                CTkMessagebox(
                    title="Error",
                    message=f"Error al abrir el archivo: {e}",
                    icon="cancel",
                )
                return

                # for idx, i in enumerate(
                #     hoja_pesas.iter_rows(min_row=2, max_col=15, values_only=True)
                # ):
                #     if i[2] is not None:  # Validamos Columna C (ID)
                #         a = str(i[0]).strip()
                #         b = str(i[1]).strip() if i[1] is not None else ""
                #         c = str(i[2]).strip() if i[2] is not None else ""
                #         e = str(i[4]).strip() if i[4] is not None else ""
                #         f = str(i[5]).strip() if i[5] is not None else ""
                #         g = str(i[6]).strip() if i[6] is not None else ""
                #         h = str(i[7]).strip() if i[7] is not None else ""
                #         j = str(i[9]).strip() if i[9] is not None else ""
                #         k = str(i[10]).strip() if i[10] is not None else ""

                # Nombre único para el checklist (con índice para evitar duplicados)
            #                 pesa_key = f"{idx+1}) {c} | {f} | {j}{k}"

            #                 # Guardamos todos los atributos en un diccionario interno
            #                 self.pesas[pesa_key] = {
            #                     "key": a,
            #                     "magnitud": b,
            #                     "id": c,
            #                     "modelo": e,
            #                     "serie": f,
            #                     "juego": g,
            #                     "id_pesa": h,
            #                     "nominal": j,
            #                     "unidad": k,
            #                 }
            #                 self.lista_pesas_completa.append(pesa_key)

            #     CTkMessagebox(
            #         title="Exito",
            #         message="Modelos cargados correctamente.",
            #         icon="check",
            #     )

            # except Exception as e:
            #     CTkMessagebox(
            #         title="Error",
            #         message=f"Error al abrir el archivo: {e}",
            #         icon="cancel",
            #     )
            #     return

    def mostrar_selector_pesas(self):
        """Filtra el diccionario de pesas y muestra en la ventana emergente solo los patrones de calibración (sustitucion == 'no')"""
        diccionario_pesas = getattr(self, "pesas", {})

        lista_patrones = [
            llave
            for llave, datos in diccionario_pesas.items()
            if datos.get("sustitucion") == "no"
        ]

        if lista_patrones:
            Ventana_Pesas(self, lista_patrones, self.actualizar_tabla_pesas)
        else:
            CTkMessagebox(
                title="Sin registros",
                message="No se encontraron elementos marcados como patrones ('no') en la columna L del archivo cargado.",
                icon="warning",
            )

    def actualizar_tabla_pesas(self, seleccionadas):
        if not hasattr(self, "llaves_actuales") or self.llaves_actuales is None:
            self.llaves_actuales = []

        for i in seleccionadas:
            if i not in self.llaves_actuales:
                self.llaves_actuales.append(i)

        for j in self.tabla_sel.get_children():
            self.tabla_sel.delete(j)

        suma_total = 0.0

        for k, l in enumerate(self.llaves_actuales, start=1):
            dato = self.pesas.get(l)
            if dato:
                self.tabla_sel.insert(
                    "",
                    "end",
                    values=(
                        k,
                        dato.get("id", "---"),
                        dato.get("serie", "---"),
                        dato.get("nominal", 0),
                        dato.get("unidad", "kg"),
                    ),
                )

                try:
                    val = float(dato.get("nominal", 0))
                    unidad = str(dato.get("unidad", "kg")).lower().strip()

                    if unidad in ["g", "gramos", "g."]:
                        suma_total += val / 1000.0
                    else:
                        suma_total += val
                except (ValueError, AttributeError):
                    pass

            self.label_suma_nominal.configure(text=f"Total: {suma_total:.2f} kg")

    def seleccionar_puerto_higro(self, puerto):
        if conectar_higrometro(puerto):
            self.hilo_higro = threading.Thread(target=read_hidro, daemon=True)
            self.hilo_higro.start()
            CTkMessagebox(
                title="Higrometro", message=f"conexion con {puerto}", icon="check"
            )
        else:
            print("mal higro")
        # if puerto == "Sin puerto":
        #     return

        # if conectar_higrometro(puerto):
        #     self.tread_lectura_h = threading.Thread(target=read_hidro, daemon=True)
        #     self.tread_lectura_h.start()
        #     CTkMessagebox(title="Éxito", message="Higrómetro conectado.", icon="check")

        # else:
        #     CTkMessagebox(
        #         title="Error",
        #         message="No se pudo conectar el higrómetro.",
        #         icon="cancel",
        #     )

    def seleccionar_puerto_bar(self, puerto):
        # if puerto == "--":
        #     return

        if conectar_barometro(puerto):
            self.hilo_bar = threading.Thread(target=read_bar, daemon=True)
            self.hilo_bar.start()
            CTkMessagebox(
                title="Barometro",
                message=f"Barometro conectado a {puerto}",
                icon="check",
            )
        else:
            CTkMessagebox(title="Barometro", message="Sin conexion", icon="cancel")
        # conexion = conectar_barometro(puerto)
        # if conexion:
        #     CTkMessagebox(title="Éxito", message="Barómetro conectado.", icon="check")
        #     self.tread_lectura_b = threading.Thread(target=read_bar, daemon=True)
        #     self.tread_lectura_b.start()
        # else:
        #     CTkMessagebox(
        #         title="Error",
        #         message="No se pudo conectar el barómetro.",
        #         icon="cancel",
        #     )

    def registrar_medicion(self):
        lectura = self.shell.peso_estable()

        if len(lectura) >= 3:
            peso_estable = int(lectura[2])
            id_fila = len(self.tabla.get_children()) + 1
            self.tabla.insert(
                "", "end", text=str(id_fila), values=(id_fila, f"{peso_estable}")
            )
            self.tabla.yview_moveto(1)

            self.desviacion()
        else:
            CTkMessagebox(
                title="Error",
                message="No se pudo obtener el peso estable.",
                icon="cancel",
            )

    def eliminar_ultimo(self):
        registro = self.tabla_sel.get_children()
        if not registro:
            CTkMessagebox(
                title="Error",
                message="No hay registros para eliminar.",
                icon="cancel",
            )
            return
        ultimo_id = registro[-1]
        self.tabla_sel.delete(ultimo_id)
        self.actualizar_suma_total()

    def actualizar_suma_total(self):
        suma = 0
        for i in self.tabla_sel.get_children():
            valor = self.tabla_sel.item(i)["values"]
            try:
                val = float(valor[3])
                unidad = str(valor[4]).lower().strip()
                if unidad in ["g", "gramos"]:
                    suma += val / 1000
            except (ValueError, IndexError):
                pass
        self.label_suma_nominal.configure(text=f"Total: {suma:.2f} kg")

    def registrar_CI(self):
        T = self.label_temperatura.cget("text")
        H = self.label_higrometro.cget("text")
        P = self.label_presion.cget("text")

        t_inst = self.combo_temperatura.get()
        h_inst = self.combo_higrometro.get()
        p_inst = self.combo_presion.get()

        self.condiciones_iniciales = {
            "Temperatura": {"Valor": T, "Instrumento": t_inst},
            "Humedad": {"Valor": H, "Instrumento": h_inst},
            "Presion": {"Valor": P, "Instrumento": p_inst},
            "Hora": datetime.now().strftime("%H:%M:%S"),
        }
        self.label_barI.configure(text=P, text_color="#158A30")
        self.label_higI.configure(text=H, text_color="#158A30")
        self.label_temI.configure(text=T, text_color="#158A30")

        self.btn_CI.configure(fg_color="#158A30")

    def registrar_CF(self):
        T = self.label_temperatura.cget("text")
        H = self.label_higrometro.cget("text")
        P = self.label_presion.cget("text")

        t_inst = self.combo_temperatura.get()
        h_inst = self.combo_higrometro.get()
        p_inst = self.combo_presion.get()

        self.condiciones_finales = {
            "Temperatura": {"Valor": T, "Instrumento": t_inst},
            "Humedad": {"Valor": H, "Instrumento": h_inst},
            "Presion": {"Valor": P, "Instrumento": p_inst},
            "Hora": datetime.now().strftime("%H:%M:%S"),
        }
        self.label_barF.configure(text=P, text_color="#D95E4A")
        self.label_higF.configure(text=H, text_color="#D95E4A")
        self.label_temF.configure(text=T, text_color="#D95E4A")

        self.btn_CF.configure(fg_color="#D95E4A")

    def eliminar_peso(self):
        registro = self.tabla.get_children()
        if not registro:
            CTkMessagebox(title="Error", message="Sin registros", icon="cancel")
            return
        ultimo_id = registro[-1]
        self.tabla.delete(ultimo_id)

    def guardar_xlsx(self):

        hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        nombre_id = f"Ejercicio_Calibracion_{hora}.xlsx"

        archivo = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=nombre_id,
            title="Guardar reporte",
        )
        if not archivo:
            return
        try:
            if os.path.exists(archivo):
                wb = openpyxl.load_workbook(archivo)
            else:
                wb = openpyxl.Workbook()
                std = (
                    wb.get_sheet_by_name("Sheet") if "Sheet" in wb.sheetnames else None
                )
                if std:
                    wb.remove(std)

            nombre = "Repetibilidad"
            contador = 1
            nuevo_nombre = nombre

            while nuevo_nombre in wb.sheetnames:
                contador += 1
                nuevo_nombre = f"{nombre} {contador}"

            hoja = wb.create_sheet(title=nuevo_nombre)

            color_encabezado = PatternFill(
                start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
            )
            fuente_negrita = Font(bold=True)

            def agregar_titulo(texto):
                hoja.append([texto])
                celda = hoja.cell(row=hoja.max_row, column=1)
                celda.fill = color_encabezado
                celda.font = fuente_negrita

            # -----INDICACIONES----
            agregar_titulo("Indicacion")
            hoja.append(["ID", "Valor"])
            for item in self.tabla.get_children():
                indice = self.tabla.item(item)["text"]
                valores = self.tabla.item(item)["values"]
                hoja.append([int(indice), float(valores[1])])
            hoja.append([])

            # -------Pesas------
            agregar_titulo("Pesas")
            hoja.append(
                [
                    "Key",
                    "Magnitud",
                    "Identificación",
                    "Modelo",
                    "Serie",
                    "Juego",
                    "Id_pesa",
                    "Nominal",
                    "Unidad",
                ]
            )
            if hasattr(self, "llaves_actuales"):
                for llave in self.llaves_actuales:
                    d = self.pesas.get(llave)
                    if d:
                        hoja.append(
                            [
                                d["key"],
                                d["magnitud"],
                                d["id"],
                                d["modelo"],
                                d["serie"],
                                d["juego"],
                                d["id_pesa"],
                                d["nominal"],
                                d["unidad"],
                            ]
                        )
            hoja.append([])

            # ---MEDIDORES---
            agregar_titulo("Medidores")
            combos = {
                "Presión": self.combo_presion.get(),
                "Humedad": self.combo_higrometro.get(),
                "Temperatura": self.combo_temperatura.get(),
            }
            for tipo, seleccion in combos.items():
                hoja.append([tipo, seleccion])
            hoja.append([])

            # ----cONDICONES AMBIENTALES----
            agregar_titulo("Condiciones_Ambientales")
            if hasattr(self, "condiciones_iniciales"):
                ci = self.condiciones_iniciales
                for mag in ["Presion", "Humedad", "Temperatura"]:
                    hoja.append(
                        [
                            "CI",
                            mag,
                            ci[mag]["Valor"],
                            # ci[mag]["Instrumento"],
                            ci["Hora"],
                        ]
                    )

            if hasattr(self, "condiciones_finales"):
                cf = self.condiciones_finales
                for mag in ["Presion", "Humedad", "Temperatura"]:
                    hoja.append(
                        [
                            "CF",
                            mag,
                            cf[mag]["Valor"],
                            # cf[mag]["Instrumento"],
                            cf["Hora"],
                        ]
                    )

            wb.save(archivo)
            CTkMessagebox(
                title="Guardar", message="Reporte guardado con éxito.", icon="check"
            )

        except Exception as e:
            CTkMessagebox(
                title="Error", message=f"No se pudo guardar: {e}", icon="cancel"
            )

    def abrir_xlsx(self):
        ruta = askopenfilename(
            title="Seleccionar archivo", filetypes=[("Excel files", "*.xlsx")]
        )
        if not ruta:
            return
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            hojas_repetibilidad = [j for j in wb.sheetnames if "Repetibilidad" in j]

            if not hojas_repetibilidad:
                CTkMessagebox(
                    title="Error",
                    message="No se encontraron hojas de repetibilidad.",
                    icon="cancel",
                )
                return
            if len(hojas_repetibilidad) > 1:
                dialogo = ctk.CTkInputDialog(
                    text=f'Seleccion de hoja:\n ({" , ".join(hojas_repetibilidad)})',
                    title="Seleccionar hoja",
                )

                nombre_hoja = dialogo.get_input()
                if nombre_hoja not in wb.sheetnames:
                    CTkMessagebox(
                        title="Error", message="Hoja no encontrada.", icon="cancel"
                    )
                    return
            else:
                nombre_hoja = hojas_repetibilidad[0]
            hoja = wb[nombre_hoja]
            for item in self.tabla.get_children():
                self.tabla.delete(item)
            for item in self.tabla_sel.get_children():
                self.tabla_sel.delete(item)

            seccion = None
            llaves_pesas_a_cargar = []

            for fila in hoja.iter_rows(values_only=True):
                if not fila or all(c is None for c in fila):
                    seccion = None
                    continue

                primer_valor = str(fila[0]).strip()

                if primer_valor == "Indicacion":
                    seccion = "MEDICIONES"
                    continue
                elif primer_valor == "Pesas":
                    seccion = "PESAS"
                    continue
                elif primer_valor == "Medidores":
                    seccion = "MEDIDORES"
                    continue
                elif primer_valor == "Condiciones_Ambientales":
                    seccion = "CONDICIONES"
                    continue

                if seccion == "MEDICIONES" and primer_valor != "ID":
                    try:
                        self.tabla.insert(
                            "",
                            "end",
                            text=primer_valor,
                            values=(primer_valor, f"{int(fila[1])}"),
                        )
                    except:
                        continue

                elif seccion == "PESAS" and primer_valor != "Key":
                    llave_excel = primer_valor
                    if llave_excel in self.lista_pesas_completa:
                        llaves_pesas_a_cargar.append(llave_excel)
                    else:
                        id_e = str(fila[2]).strip()
                        ser_e = str(fila[4]).strip()
                        for llave_m in self.lista_pesas_completa:
                            if id_e in llave_m and ser_e in llave_m:
                                llaves_pesas_a_cargar.append(llave_m)
                                break

                elif seccion == "MEDIDORES":
                    tipo, seleccion = primer_valor, str(fila[1])
                    if seleccion and seleccion != "None":
                        if "Presión" in tipo:
                            self.combo_presion.set(seleccion)
                        elif "Humedad" in tipo:
                            self.combo_higrometro.set(seleccion)
                        elif "Temperatura" in tipo:
                            self.combo_temperatura.set(seleccion)

                elif seccion == "CONDICIONES":
                    etapa, mag, valor = fila[0], fila[1], str(fila[2])
                    if etapa == "CI":
                        if "Presion" in mag:
                            self.label_barI.configure(text=valor, text_color="#158A30")
                        if "Humedad" in mag:
                            self.label_higI.configure(text=valor, text_color="#158A30")
                        if "Temperatura" in mag:
                            self.label_temI.configure(text=valor, text_color="#158A30")
                    elif etapa == "CF":
                        if "Presion" in mag:
                            self.label_barF.configure(text=valor, text_color="#D95E4A")
                        if "Humedad" in mag:
                            self.label_higF.configure(text=valor, text_color="#D95E4A")
                        if "Temperatura" in mag:
                            self.label_temF.configure(text=valor, text_color="#D95E4A")
            if llaves_pesas_a_cargar:
                self.actualizar_tabla_pesas(llaves_pesas_a_cargar)

            self.desviacion()
            CTkMessagebox(
                title="Éxito", message="Reporte cargado correctamente", icon="check"
            )

        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"No se pudo cargar el archivo: {e}",
                icon="cancel",
            )

    def desviacion(self):
        items = self.tabla.get_children()
        registros = []
        for item in items:
            try:
                valor = self.tabla.item(item)["values"][1]
                registros.append(valor)
            except (ValueError, IndexError):
                continue

        if len(registros) < 2:
            self.label_desviacion.configure(text="S: N/A")
            return

        df_serie = pd.Series(registros)

        std_dev = df_serie.std()
        self.label_desviacion.configure(text=f"S: {std_dev:.4f} kg")

    def nueva_prueba(self):
        for item in self.tabla.get_children():
            self.tabla.delete(item)
        for item in self.tabla_sel.get_children():
            self.tabla_sel.delete(item)

        self.label_suma_nominal.configure()  # text="Suma Nominal: 0 kg")
        self.label_barI.configure()  # text="--", text_color="black")
        self.label_higI.configure()  # text="--", text_color="black")
        self.label_temI.configure()  # text="--", text_color="black")
        self.label_barF.configure()  # text="---", text_color="black")
        self.label_higF.configure()  # text="---", text_color="black")
        self.label_temF.configure()  # text="---", text_color="black")
        self.btn_CI.configure()  # )
        self.btn_CF.configure()  # fg_color="#4CAF50")

    def mostrar_selector_sustitucion(self):
        diccionario_pesas = getattr(self, "pesas", {})

        lista_sustitucion = [
            llave
            for llave, datos in diccionario_pesas.items()
            if datos.get("sustitucion") == "si"
        ]

        if lista_sustitucion:
            Ventana_Pesas(self, lista_sustitucion, self.actualizar_tabla_pesas)
        else:
            CTkMessagebox(
                title="Sin registros",
                message="No se encontraron elementos marcados como sustitución archivo cargado.",
                icon="warning",
            )


class Ventana_Pesas(ctk.CTkToplevel):
    def __init__(self, master, lista_pesas, callback_guardar):
        super().__init__(master)
        self.title("Selección de Pesas")
        self.geometry("450x650+200+100")
        self.grab_set()

        self.callback_guardar = callback_guardar
        self.nombres_indexados = lista_pesas
        self.vars_dict = {}

        self.lista_vars = []

        # ---Ranfo---
        self.frame_rango = ctk.CTkFrame(self)
        self.frame_rango.pack(fill="x", padx=10, pady=10)

        ctk.CTkLabel(
            self.frame_rango, text="Intervalo (ej. 1-5, 8):", font=("Cambria", 12)
        ).pack(side="left", padx=5)

        self.entry_rango = ctk.CTkEntry(
            self.frame_rango, placeholder_text="1-5 o 1,3", width=100, justify="center"
        )
        self.entry_rango.pack(side="left", padx=5)

        self.btn_aplicar = ctk.CTkButton(
            self.frame_rango,
            text="Seleccionar",
            width=80,
            command=self.seleccionar_por_rango,
        )
        self.btn_aplicar.pack(side="left", padx=5)

        self.scroll = ctk.CTkScrollableFrame(self)
        self.scroll.pack(fill="both", expand=True, padx=10, pady=5)

        for i, p in enumerate(self.nombres_indexados, start=1):
            v = tk.BooleanVar(self, value=False)
            texto_visual = f"{i}) {p}"
            cb = ctk.CTkCheckBox(self.scroll, text=texto_visual, variable=v)
            cb.pack(pady=2, anchor="w")

            self.vars_dict[p] = v
            self.lista_vars.append(v)

        self.boton_confirmar = ctk.CTkButton(
            self, text="Confirmar Selección", fg_color="green", command=self.enviar
        )
        self.boton_confirmar.pack(pady=10)

    def seleccionar_por_rango(self):
        entrada = self.entry_rango.get().replace(" ", "")
        if not entrada:
            return

        try:
            partes = entrada.split(",")
            for parte in partes:
                if "-" in parte:
                    inicio, fin = map(int, parte.split("-"))
                    for i in range(inicio, fin + 1):
                        if 1 <= i <= len(self.lista_vars):
                            self.lista_vars[i - 1].set(True)
                else:
                    idx = int(parte)
                    if 1 <= idx <= len(self.lista_vars):
                        self.lista_vars[idx - 1].set(True)

            self.entry_rango.delete(0, "end")
        except ValueError:
            print("Rango invalido")

    def enviar(self):
        seleccion = [n for n, v in self.vars_dict.items() if v.get()]
        self.callback_guardar(seleccion)
        self.destroy()


class Excentricidad(ctk.CTkToplevel):
    def __init__(
        self,
        master,
        shell: Shell,
        lista_instrumentos,
        ruta_excel_trabajo=None,
        *args,
        **kwargs,
    ):
        super().__init__(master, *args, **kwargs)
        self.shell = shell
        self.lista_instrumentos = lista_instrumentos
        self.ruta_excel_trabajo = ruta_excel_trabajo
        self.title("Excentricidad")
        self.geometry("1200x800+190+20")
        self.grab_set()
        self.menu()
        self.archivo_calibracion_actual = None
        self.widgets()
        self.sentido_actual = "IDA"
        self.actualizar_peso()

    def widgets(self):
        self.frame_repetibilidad = ctk.CTkFrame(self, width=800, height=700)
        self.frame_repetibilidad.pack(fill="both", expand=True, padx=5, pady=5)
        self.frame_repetibilidad.pack_propagate(False)

        self.frame_peso = ctk.CTkFrame(self.frame_repetibilidad, width=400, height=150)
        self.frame_peso.pack(anchor="n", padx=10, pady=10, fill="x")
        self.frame_peso.pack_propagate(False)

        self.frame_peso_instantaneo = ctk.CTkFrame(
            self.frame_peso, width=450, height=150
        )
        self.frame_peso_instantaneo.pack(side="left", padx=10, pady=10)
        self.frame_peso_instantaneo.pack_propagate(False)

        self.label_unidad = ctk.CTkLabel(
            self.frame_peso_instantaneo, text="kg", font=("Cambria", 50)
        )
        self.label_unidad.pack(side="right", padx=20, pady=5)

        self.frame_peso_al_momento = ctk.CTkFrame(
            self.frame_peso_instantaneo, width=450, height=150
        )
        self.frame_peso_al_momento.pack(side="left", padx=10, pady=10)
        self.frame_peso_al_momento.pack_propagate(False)

        self.label_peso = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 70)
        )
        self.label_peso.pack(padx=20, pady=5, expand=True)

        self.label_tara = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 12)
        )
        self.label_tara.place(x=8, y=8)

        self.label_neto = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 12)
        )
        self.label_neto.place(x=8, y=79)

        # ---------BOTONES-------------------
        imagen_tarar = Image.open(ruta(os.path.join("Icon", "aumentar.png")))
        icon_tarar = ctk.CTkImage(
            light_image=imagen_tarar, dark_image=imagen_tarar, size=(30, 30)
        )

        self.button_tarar = ctk.CTkButton(
            self.frame_peso,
            text="Tara",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.tara,
            image=icon_tarar,
            compound="top",
        )
        self.button_tarar.pack(side="left", expand=True)
        # ---------------BOTON ZERO-------------
        imagen_zero = Image.open(ruta(os.path.join("Icon", "cero.png")))
        icon_zero = ctk.CTkImage(
            light_image=imagen_zero, dark_image=imagen_zero, size=(30, 30)
        )

        self.button_zero = ctk.CTkButton(
            self.frame_peso,
            text="Zero",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.zero,
            image=icon_zero,
            compound="top",
        )
        self.button_zero.pack(side="left", expand=True)

        # ---------------BOTON QUITAR TARAR-------------
        imagen_q_tara = Image.open(ruta(os.path.join("Icon", "perdida-peso.png")))
        icon_q_tara = ctk.CTkImage(
            light_image=imagen_q_tara, dark_image=imagen_q_tara, size=(30, 30)
        )
        self.button_quitar_tara = ctk.CTkButton(
            self.frame_peso,
            text="Eliminar Tara",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.quitar_tara,
            image=icon_q_tara,
            compound="top",
        )
        self.button_quitar_tara.pack(side="left", expand=True)

        # ----------------Condiciones Ambientales-----------------
        self.frame_condiciones = ctk.CTkFrame(
            self.frame_repetibilidad, width=400, height=150
        )
        self.frame_condiciones.pack(anchor="s", padx=10, pady=5, fill="x")
        self.frame_condiciones.pack_propagate(False)

        self.frame_condicion1 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion1.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion1.pack_propagate(False)

        self.label_condiciones = ctk.CTkLabel(
            self.frame_condicion1,
            text="Presión",
            font=("Cambria", 20),
        )
        self.label_condiciones.pack(anchor="n", padx=2, pady=5, expand=True)

        self.frame_condicion2 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion2.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion2.pack_propagate(False)

        self.label_condiciones2 = ctk.CTkLabel(
            self.frame_condicion2,
            text="Humedad",
            font=("Cambria", 20),
        )
        self.label_condiciones2.pack(anchor="n", padx=2, pady=5, expand=True)

        self.frame_condicion3 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion3.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion3.pack_propagate(False)

        self.label_condiciones3 = ctk.CTkLabel(
            self.frame_condicion3,
            text="Temperatura",
            font=("Cambria", 20),
        )
        self.label_condiciones3.pack(anchor="n", padx=2, pady=5, expand=True)
        # ---------PRESION-----------------

        self.frame_presion = ctk.CTkFrame(self.frame_condicion1, width=200, height=50)
        self.frame_presion.pack(expand=True, anchor="n")
        self.frame_presion.pack_propagate(False)

        self.label_presion = ctk.CTkLabel(
            self.frame_presion,
            text="---",
            font=("Cambria", 30),
        )
        self.label_presion.pack(expand=True)

        self.combo_presion = ctk.CTkOptionMenu(
            master=self.frame_condicion1,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_presion.pack(expand=True)
        self.combo_presion.set("Modelos")

        # -------------HIGROMETRO-----------------

        self.frame_higrometro = ctk.CTkFrame(
            self.frame_condicion2, width=200, height=50
        )
        self.frame_higrometro.pack(expand=True, anchor="n")
        self.frame_higrometro.pack_propagate(False)

        self.label_higrometro = ctk.CTkLabel(
            self.frame_higrometro,
            text="---",
            font=("Cambria", 30),
        )
        self.label_higrometro.pack(expand=True)
        self.combo_higrometro = ctk.CTkOptionMenu(
            master=self.frame_condicion2,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_higrometro.pack(expand=True)
        self.combo_higrometro.set("Modelos")

        # --------------Temperatura-----------------

        self.frame_temperatura = ctk.CTkFrame(
            self.frame_condicion3, width=200, height=50
        )
        self.frame_temperatura.pack(expand=True, anchor="n")
        self.frame_temperatura.pack_propagate(False)

        self.label_temperatura = ctk.CTkLabel(
            self.frame_temperatura, text="---", font=("Cambria", 30)
        )
        self.label_temperatura.pack(expand=True)
        self.combo_temperatura = ctk.CTkOptionMenu(
            master=self.frame_condicion3,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_temperatura.pack(expand=True)
        self.combo_temperatura.set("Modelos")

        # --------------------------------
        self.frame_registro = ctk.CTkFrame(
            self.frame_repetibilidad, width=400, height=420
        )
        self.frame_registro.pack(anchor="s", padx=10, pady=5, fill="x")
        self.frame_registro.pack_propagate(False)
        # ----------------------------------
        # -------Tabla de registro/Entry------
        self.tabla = ttk.Treeview(
            self.frame_registro,
            columns=(
                "sentido",
                "seccion",
                "indicacion",
            ),
            show="headings",
            height=10,
        )

        self.tabla.heading("sentido", text="Sentido")
        self.tabla.heading("seccion", text="Sección")
        self.tabla.heading("indicacion", text="Indicación")

        self.tabla.column("sentido", width=60, anchor="center")
        self.tabla.column("seccion", width=60, anchor="center")
        self.tabla.column("indicacion", width=90, anchor="center")
        self.tabla.pack(side="left", padx=10, pady=20, anchor="n")

        # ------tabla de prmedios------
        self.tabla_promedios = ttk.Treeview(
            self.frame_registro,
            columns=("seccion_p", "promedio"),
            show="headings",
            height=10,
        )

        self.tabla_promedios.heading("seccion_p", text="Sección")
        self.tabla_promedios.heading("promedio", text="Promedio")

        self.tabla_promedios.column("seccion_p", width=90, anchor="center")
        self.tabla_promedios.column("promedio", width=90, anchor="center")

        self.tabla_promedios.pack(side="left", padx=15, pady=20, anchor="n")

        # ----------frame de botones excentricidad----------
        self.frame_botones_EX = ctk.CTkFrame(self.frame_registro, width=390, height=130)
        self.frame_botones_EX.pack(side="left", padx=10, pady=20, anchor="n")
        self.frame_botones_EX.pack_propagate(False)

        self.entry_secciones = ctk.CTkEntry(
            self.frame_botones_EX,
            placeholder_text="Celdas",
            width=100,
            justify="center",
            border_color="white",
        )
        self.entry_secciones.grid(row=0, column=0, padx=5, pady=10)

        # self.entry_secciones.pack(side="top", padx=5, pady=10, anchor="n")
        # self.entry_secciones.insert(0,"1")

        imagen_generar = Image.open(ruta(os.path.join("Icon", "idea.png")))
        icon_generar = ctk.CTkImage(
            light_image=imagen_generar, dark_image=imagen_generar, size=(30, 30)
        )

        self.btn_generar = ctk.CTkButton(
            self.frame_botones_EX,
            text="Generar",
            command=self.preparar_tabla_dinamica,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_generar,
        )
        self.btn_generar.grid(row=1, column=0, padx=5, pady=10)
        # self.btn_generar.pack(side="top", padx=5, pady=5, anchor="n")

        # imagen_sentido = Image.open(ruta(os.path.join("Icon", "de-dos-vias.png")))
        # icon_sentido = ctk.CTkImage(
        #     light_image=imagen_sentido, dark_image=imagen_sentido, size=(30, 30)
        # )
        # self.btn_sentido = ctk.CTkButton(
        #     self.frame_botones_EX,
        #     text="Sentido",
        #     command=self.cambiar_sentido,
        #     fg_color="#158A30",
        #     image=icon_sentido,
        # )
        # self.btn_sentido.grid(row=2, column=0, padx=5, pady=10)

        # self.frame_btn_principal = ctk.CTkFrame(
        #     self.frame_registro, width=200, height=250
        # )

        # # self.frame_btn_principal.place(x=450, y=180)
        # # self.frame_btn_principal.pack_propagate(False)

        imagen_indicacion = Image.open(ruta(os.path.join("Icon", "boton-agregar.png")))
        icon_indicacion = ctk.CTkImage(
            light_image=imagen_indicacion, dark_image=imagen_indicacion, size=(30, 30)
        )
        self.btn_indicacion = ctk.CTkButton(
            self.frame_botones_EX,
            text="Registrar Peso",
            command=self.registrar_peso,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_indicacion,
        )
        self.btn_indicacion.grid(row=3, column=0, padx=5, pady=10)

        # ---------GUARDAR---------
        imagen_btn_guardar = Image.open(
            ruta(os.path.join("Icon", "guardar-el-archivo.png"))
        )
        icon_btn_guardar = ctk.CTkImage(
            light_image=imagen_btn_guardar, dark_image=imagen_btn_guardar, size=(30, 30)
        )
        self.btn_guardar = ctk.CTkButton(
            self.frame_botones_EX,
            text="Guardar",
            command=self.guardar_en_excel_existente,
            fg_color="#4CAF50",
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_btn_guardar,
        )
        self.btn_guardar.grid(row=2, column=0, padx=5, pady=10)

        # --------BOTONES DE PESAS SUSTITUCION O NADA----

        imagen_pesas = Image.open(ruta(os.path.join("Icon", "escoger.png")))
        icon_pesas = ctk.CTkImage(
            light_image=imagen_pesas, dark_image=imagen_pesas, size=(30, 30)
        )

        self.btn_pesas = ctk.CTkButton(
            self.frame_botones_EX,
            text="Pesas",
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.mostrar_selector_pesas,
            image=icon_pesas,
        )
        self.btn_pesas.grid(row=0, column=1, padx=5, pady=10)

        self.btn_sustitucion = ctk.CTkButton(
            self.frame_botones_EX,
            text="Sustitución",
            command=self.mostrar_selector_sustitucion,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_pesas,
        )
        self.btn_sustitucion.grid(row=1, column=1, padx=5, pady=10)

        # ----------Tabla de pesos -----------

        self.tabla_sel = ttk.Treeview(
            self.frame_registro,
            columns=("#", "ID", "serie", "valor", "unidad"),
            show="headings",
            height=8,
        )
        self.tabla_sel.heading("#", text="#")
        self.tabla_sel.heading("ID", text="Identificación")
        self.tabla_sel.heading("serie", text="Serie")
        self.tabla_sel.heading("valor", text="V nominal")
        self.tabla_sel.heading("unidad", text="Unidad")

        self.tabla_sel.column("#", width=30, anchor="center")
        self.tabla_sel.column("ID", width=90, anchor="center")
        self.tabla_sel.column("serie", width=90, anchor="center")
        self.tabla_sel.column("valor", width=90, anchor="center")
        self.tabla_sel.column("unidad", width=70, anchor="center")
        self.tabla_sel.pack(side="left", padx=5, pady=20, anchor="n")

        self.label_suma_nominal = ctk.CTkLabel(
            self.frame_registro, text="Suma Nominal: 0 kg", font=("Cambria", 12)
        )
        self.label_suma_nominal.place(x=990, y=250)

        # ------Condiciones ----------------
        # _------------------condiciones----
        self.frame_R_condiciones = ctk.CTkFrame(
            self.frame_registro, width=450, height=45
        )
        self.frame_R_condiciones.place(x=380, y=320)
        self.frame_R_condiciones.pack_propagate(False)

        self.frame_r_bar = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_bar.pack(side="left", expand=True, padx=10)
        self.frame_r_bar.pack_propagate(False)

        self.label_barI = ctk.CTkLabel(
            self.frame_r_bar, text="--", font=("Cambria", 15)
        )
        self.label_barI.pack(expand=True)

        self.frame_r_pres = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_pres.pack(side="left", expand=True)
        self.frame_r_pres.pack_propagate(False)

        self.label_higI = ctk.CTkLabel(
            self.frame_r_pres, text="--", font=("Cambria", 15)
        )
        self.label_higI.pack(expand=True)

        self.frame_r_tem = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_tem.pack(side="left", expand=True)
        self.frame_r_tem.pack_propagate(False)

        self.label_temI = ctk.CTkLabel(
            self.frame_r_tem, text="--", font=("Cambria", 15)
        )
        self.label_temI.pack(expand=True)

        # ---------condiciones finales ------

        self.frame_cf = ctk.CTkFrame(self.frame_registro, width=450, height=45)
        self.frame_cf.place(x=380, y=370)
        self.frame_cf.pack_propagate(False)

        self.frame_barF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_barF.pack(side="left", expand=True, padx=10)
        self.frame_barF.pack_propagate(False)

        self.label_barF = ctk.CTkLabel(
            self.frame_barF, text="---", font=("Cambria", 15)
        )
        self.label_barF.pack(expand=True)

        self.frame_higF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_higF.pack(side="left", expand=True)
        self.frame_higF.pack_propagate(False)
        self.label_higF = ctk.CTkLabel(
            self.frame_higF, text="---", font=("Cambria", 15)
        )
        self.label_higF.pack(expand=True)

        self.frame_temF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_temF.pack(side="left", expand=True)
        self.frame_temF.pack_propagate(False)

        self.label_temF = ctk.CTkLabel(
            self.frame_temF, text="---", font=("Cambria", 15)
        )
        self.label_temF.pack(expand=True)

        # ------Btonoes de condiones------
        imagen_C = Image.open(ruta(os.path.join("Icon", "escribir.png")))
        icon_C = ctk.CTkImage(light_image=imagen_C, dark_image=imagen_C, size=(30, 30))
        self.btn_CI = ctk.CTkButton(
            self.frame_botones_EX,
            text="Registrar CI",
            width=150,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            command=self.registrar_CI,
            image=icon_C,
        )

        self.btn_CI.grid(row=2, column=1, padx=5, pady=10)

        # self.btn_CI.pack(side="top", expand=True, padx=10, pady=10)

        self.btn_CF = ctk.CTkButton(
            self.frame_botones_EX,
            text="Registrar CF",
            width=150,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            command=self.registrar_CF,
            image=icon_C,
        )

        self.btn_CF.grid(row=3, column=1, padx=5, pady=10)

        # --------Botones de borrar registro o borrar todo-------

        imagen_borrar = Image.open(ruta(os.path.join("Icon", "borrar.png")))
        icon_borrar = ctk.CTkImage(
            light_image=imagen_borrar, dark_image=imagen_borrar, size=(30, 30)
        )

        self.btn_borrar_pesa = ctk.CTkButton(
            self.frame_botones_EX,
            text="Borrar Pesa",
            command=self.eliminar_ultima_pesa,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_borrar,
        )
        self.btn_borrar_pesa.grid(row=4, column=0, padx=5, pady=10)

        imagen_borrar_indicacion = Image.open(
            ruta(os.path.join("Icon", "borrarRojo.png"))
        )
        icon_borrar_indicacion = ctk.CTkImage(
            light_image=imagen_borrar_indicacion,
            dark_image=imagen_borrar_indicacion,
            size=(30, 30),
        )

        self.btn_borrar_indicacion = ctk.CTkButton(
            self.frame_botones_EX,
            text="Borrar indicacion",
            command=self.eliminar_ultima_indicacion,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_borrar_indicacion,
        )
        self.btn_borrar_indicacion.grid(row=4, column=1, padx=5, pady=10)

    def menu(self):
        self.menu = CTkMenuBar(master=self)
        self.b1 = self.menu.add_cascade("Archivo", command=self.abrir_modelos)
        self.b2_puerto_bar = self.menu.add_cascade("Barometro")
        self.dropdown_puerto_bar = CustomDropdownMenu(widget=self.b2_puerto_bar)
        self.b3_puerto_higro = self.menu.add_cascade("Higrometro")
        self.dropdown_puerto_higro = CustomDropdownMenu(widget=self.b3_puerto_higro)
        self.b4 = self.menu.add_cascade(
            "Leer xlsx", command=self.master.cargar_excel_calibracion
        )
        self.b5 = self.menu.add_cascade(
            "Cargar Prueba", command=self.cargar_datos_desde_excel
        )
        self.b6 = self.menu.add_cascade("Nueva Prueba", command=self.nueva_prueba)

        puertos = serial.tools.list_ports.comports()
        if not puertos:
            self.dropdown_puerto_bar.add_command(label="No se encontraron puertos")
            self.dropdown_puerto_higro.add_command(label="No se encontraron puertos")
        else:
            for i in puertos:
                self.dropdown_puerto_bar.add_option(
                    option=i.device,
                    command=lambda puerto=i.device: self.seleccionar_puerto_bar(puerto),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                )
                self.dropdown_puerto_higro.add_option(
                    option=i.device,
                    command=lambda puerto=i.device: self.seleccionar_puerto_higro(
                        puerto
                    ),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                )

    def actualizar_peso(self):
        global Lectura_barometro, LecturaT1, LecturaT2
        try:
            val_menos = "{:.2f}".format(float(Lectura_barometro))
            self.label_presion.configure(text=f"{val_menos} Pa")
            self.label_temperatura.configure(text=f"{LecturaT1} °C")
            self.label_higrometro.configure(text=f"{LecturaH1} %")
        except Exception as e:
            print(f"error -{e}")

        lecture = self.shell.peso_instantaneo()

        if len(lecture) >= 3:
            try:
                peso_r = int(float(lecture[2]))
                self.label_peso.configure(text=str(peso_r))
                self.label_neto.configure(text=f"Neto: {peso_r} kg", text_color="white")
            except ValueError:
                pass

        if len(lecture) >= 3:
            wg = lecture[2]
            self.label_peso.configure(text=wg)

            if lecture[1] == "S":
                self.label_peso.configure(text_color="#158A30")
            else:
                self.label_peso.configure(text_color="#B52C19")

            valor_tara = self.shell.obtener_tara()
            self.label_tara.configure(text=f"Tara: {valor_tara}", text_color="#FF6021")

        self.after(500, self.actualizar_peso)

    def abrir_modelos(self):
        ruta = askopenfilename(
            title="Seleccionar archivo", filetypes=[("Archivos de Excel", "*.xlsx")]
        )

        if ruta:
            try:
                wb = openpyxl.load_workbook(ruta, data_only=True)

                if "Equipo CA" in wb.sheetnames:
                    hoja = wb["Equipo CA"]
                    self.equipos_ca = {}
                else:
                    CTkMessagebox(
                        title="Error",
                        message='La hoja "Equipo CA" no se encuentra en el archivo',
                        icon="error",
                    )
                    return

                lista_sensores = []
                for fila in hoja.iter_rows(min_row=2, max_col=10, values_only=True):
                    if fila[0] is not None:
                        a = str(fila[0]).strip()
                        b = str(fila[1]).strip() if fila[1] is not None else ""
                        c = str(fila[2]).strip() if fila[2] is not None else ""
                        d = str(fila[3]).strip() if fila[3] is not None else ""
                        e = str(fila[5]).strip() if fila[5] is not None else ""
                        h = str(fila[7]).strip() if fila[7] is not None else ""
                        sensor_info = f"{d}, {e}, {h}".strip(", ")

                        self.equipos_ca[sensor_info] = {
                            "key": a,
                            "magnitud": b,
                            "indicacion": c,
                            "marca": d,
                            "modelo": e,
                            "serie_sensor": h,
                        }

                        if sensor_info not in lista_sensores:
                            lista_sensores.append(sensor_info)
                self.lista_medidores = lista_sensores
                self.combo_presion.configure(values=lista_sensores)
                self.combo_higrometro.configure(values=lista_sensores)
                self.combo_temperatura.configure(values=lista_sensores)

                if "Equipo Pesas" in wb.sheetnames:
                    hoja_pesas = wb["Equipo Pesas"]
                    self.pesas = {}  # Diccionario maestro local
                    self.lista_pesas_completa = []

                    for idx, i in enumerate(
                        hoja_pesas.iter_rows(min_row=2, max_col=12, values_only=True)
                    ):
                        if i[2] is not None:
                            a = str(i[0]).strip()
                            b = str(i[1]).strip() if i[1] is not None else ""
                            c = str(i[2]).strip() if i[2] is not None else ""
                            e = str(i[4]).strip() if i[4] is not None else ""
                            f = str(i[5]).strip() if i[5] is not None else ""
                            g = str(i[6]).strip() if i[6] is not None else ""
                            h = str(i[7]).strip() if i[7] is not None else ""
                            j = str(i[9]).strip() if i[9] is not None else ""
                            k = str(i[10]).strip() if i[10] is not None else ""

                            sust = (
                                str(i[11]).strip().lower()
                                if i[11] is not None
                                else "no"
                            )

                            pesa_key = f"{c} | {f} | {j}{k}"
                            self.pesas[pesa_key] = {
                                "key": a,
                                "magnitud": b,
                                "id": c,
                                "modelo": e,
                                "serie": f,
                                "juego": g,
                                "id_pesa": h,
                                "nominal": j,
                                "unidad": k,
                                "sustitucion": sust,
                            }
                            self.lista_pesas_completa.append(pesa_key)

                CTkMessagebox(
                    title="Exito",
                    message="Modelos cargados correctamente.",
                    icon="check",
                )

            except Exception as e:
                CTkMessagebox(
                    title="Error",
                    message=f"Error al abrir el archivo: {e}",
                    icon="cancel",
                )
                return

    def seleccionar_puerto_higro(self, puerto):
        if conectar_higrometro(puerto):
            self.hilo_higro = threading.Thread(target=read_hidro, daemon=True)
            self.hilo_higro.start()
            CTkMessagebox(
                title="Higrometro", message=f"conexion con {puerto}", icon="check"
            )
        else:
            print("mal higro")

    def seleccionar_puerto_bar(self, puerto):

        if conectar_barometro(puerto):
            self.hilo_bar = threading.Thread(target=read_bar, daemon=True)
            self.hilo_bar.start()
            CTkMessagebox(
                title="Barometro",
                message=f"Barometro conectado a {puerto}",
                icon="check",
            )
        else:
            CTkMessagebox(title="Barometro", message="Sin conexion", icon="cancel")

    def preparar_tabla_dinamica(self):
        try:
            n_secciones = int(self.entry_secciones.get())
            msg = CTkMessagebox(
                title="Sentido de Entrada",
                message="¿Por qué lado ingresará el vehículo a la plataforma?",
                icon="question",
                option_1="Derecha (Der-Izq)",
                option_2="Izquierda (Izq-Der)",
                option_3="Cancelar",
            )
            response = msg.get()
            if response == "Cancelar" or response is None:
                return

            for i in self.tabla.get_children():
                self.tabla.delete(i)

            if response == "Derecha (Der-Izq)":
                txt_ida = "Der.-Izq."
                txt_vta = "Izq.-Der."
            else:
                txt_ida = "Izq.-Der."
                txt_vta = "Der.-Izq."

            self.sentido_actual = "IDA"
            if hasattr(self, "btn_sentido"):
                self.btn_sentido.configure(text="Sentido: IDA", fg_color="#16a085")

            if n_secciones % 2 != 0:
                seccion_centro = (n_secciones // 2) + 1
                punto_insercion_par = -1
            else:
                seccion_centro = -1
                punto_insercion_par = (n_secciones // 2) + 1

            filas_totales = []

            if response == "Derecha (Der-Izq)":
                for s in range(1, n_secciones):
                    if s == punto_insercion_par:
                        filas_totales.append((txt_ida, "Centro", "---"))
                    sec_str = "Centro" if s == seccion_centro else str(s)
                    filas_totales.append((txt_ida, sec_str, "---"))

                for s in range(n_secciones, 1, -1):
                    if s == punto_insercion_par - 1:
                        filas_totales.append((txt_vta, "Centro", "---"))
                    sec_str = "Centro" if s == seccion_centro else str(s)
                    filas_totales.append((txt_vta, sec_str, "---"))
            else:
                for s in range(n_secciones, 1, -1):
                    if s == punto_insercion_par - 1:
                        filas_totales.append((txt_ida, "Centro", "---"))
                    sec_str = "Centro" if s == seccion_centro else str(s)
                    filas_totales.append((txt_ida, sec_str, "---"))

                for s in range(1, n_secciones):
                    if s == punto_insercion_par:
                        filas_totales.append((txt_vta, "Centro", "---"))
                    sec_str = "Centro" if s == seccion_centro else str(s)
                    filas_totales.append((txt_vta, sec_str, "---"))

            for sent, sec, ind in filas_totales:
                self.tabla.insert("", "end", values=(sent, sec, ind))

        except ValueError:
            CTkMessagebox(
                title="Error", message="Número de secciones inválido", icon="cancel"
            )

    def registrar_peso(self):
        lectura = self.shell.peso_estable()
        if len(lectura) >= 3:
            peso = lectura[2]

            for item in self.tabla.get_children():
                v = self.tabla.item(item)["values"]

                if str(v[2]).strip() == "---":
                    self.tabla.item(item, values=(v[0], v[1], peso))
                    self.tabla.selection_set(item)
                    self.tabla.see(item)

                    self.calcular_promedios_secciones()
                    return

            CTkMessagebox(
                title="Aviso",
                message="Todas las indicaciones de la prueba están completas.",
                icon="info",
            )
        else:
            CTkMessagebox(title="Error", message="Peso inestable", icon="cancel")

    def cambiar_sentido(self):
        if self.sentido_actual == "IDA":
            self.sentido_actual = "VUELTA"
            self.btn_sentido.configure(text="Sentido: Izq-Der", fg_color="#2980b9")
        else:
            self.sentido_actual = "IDA"
            self.btn_sentido.configure(text="Sentido: Der-Izq", fg_color="#16a085")

        CTkMessagebox(title="Sentido", message=f"Cambiado a: {self.sentido_actual}")

    def calcular_promedios_secciones(self):
        datos_para_promediar = []

        for item in self.tabla.get_children():
            v = self.tabla.item(item)["values"]

            if (
                len(v) >= 3
                and v[1]
                and str(v[2]).strip() != "---"
                and str(v[2]).strip() != ""
            ):
                try:
                    datos_para_promediar.append(
                        {"seccion": str(v[1]).strip(), "valor": float(v[2])}
                    )
                except ValueError:
                    pass

        if not datos_para_promediar:
            for i in self.tabla_promedios.get_children():
                self.tabla_promedios.delete(i)
            return

        df = pd.DataFrame(datos_para_promediar)

        resumen = df.groupby("seccion")["valor"].mean().reset_index()

        for i in self.tabla_promedios.get_children():
            self.tabla_promedios.delete(i)

        for _, fila in resumen.iterrows():
            self.tabla_promedios.insert(
                "", "end", values=(fila["seccion"], f"{fila['valor']:.2f}")
            )

    def guardar_en_excel_existente(self):
        from datetime import datetime

        filas_tabla = self.tabla.get_children()
        if not filas_tabla:
            CTkMessagebox(
                title="Aviso",
                message="No hay datos de indicaciones en la tabla para guardar.",
                icon="warning",
            )
            return

        ruta_directa = getattr(self.master, "archivo_calibracion_actual", None)

        if not ruta_directa:
            self.guardar_como_nuevo_excel()
            return

        # if not ruta_directa:
        #     CTkMessagebox(
        #         title="Archivo Faltante",
        #         message="Cargue un archivo de calibración antes de guardar.",
        #         icon="warning",
        #     )
        #     return

        try:
            wb = openpyxl.load_workbook(ruta_directa)

            if (
                not hasattr(self, "hoja_actual_guardada")
                or self.hoja_actual_guardada is None
            ):
                if "Excentricidad" not in wb.sheetnames:
                    nombre_hoja = "Excentricidad"
                else:
                    contador = 2
                    while f"Excentricidad {contador}" in wb.sheetnames:
                        contador += 1
                    nombre_hoja = f"Excentricidad {contador}"

                self.hoja_actual_guardada = nombre_hoja
                hoja = wb.create_sheet(title=nombre_hoja)
            else:
                nombre_hoja = self.hoja_actual_guardada
                if nombre_hoja not in wb.sheetnames:
                    hoja = wb.create_sheet(title=nombre_hoja)
                else:
                    hoja = wb[nombre_hoja]
                    hoja.delete_rows(1, hoja.max_row + 1)

            hoja.views.sheetView[0].showGridLines = True

            color_encabezado = PatternFill(
                start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
            )
            fuente_negrita = Font(bold=True)

            def agregar_titulo_bloque(texto):
                hoja.append([texto])
                celda = hoja.cell(row=hoja.max_row, column=1)
                celda.fill = color_encabezado
                celda.font = fuente_negrita

            agregar_titulo_bloque("Indicacion")
            hoja.append(["Sentido", "Sección", "Valor (kg)"])

            for item in filas_tabla:
                v = self.tabla.item(item)["values"]
                val_ind = (
                    float(v[2])
                    if str(v[2]).strip() != "---" and str(v[2]).strip() != ""
                    else v[2]
                )
                hoja.append([v[0], v[1], val_ind])

            hoja.append([])

            agregar_titulo_bloque("Promedios por Sección")
            hoja.append(["Sección", "Promedio (kg)"])

            for item in self.tabla_promedios.get_children():
                vp = self.tabla_promedios.item(item)["values"]
                try:
                    hoja.append([vp[0], float(vp[1])])
                except ValueError:
                    hoja.append([vp[0], vp[1]])

            hoja.append([])

            agregar_titulo_bloque("Pesas")
            hoja.append(
                [
                    "Key",
                    "Magnitud",
                    "Identificación",
                    "Modelo",
                    "Serie",
                    "Juego",
                    "Id_pesa",
                    "Nominal",
                    "Unidad",
                ]
            )

            if hasattr(self, "llaves_actuales") and self.llaves_actuales:
                diccionario_local_pesas = getattr(self, "pesas", {})
                for llave in self.llaves_actuales:
                    d = diccionario_local_pesas.get(llave)
                    if d:
                        hoja.append(
                            [
                                d.get("key", ""),
                                d.get("magnitud", ""),
                                d.get("id", ""),
                                d.get("modelo", ""),
                                d.get("serie", ""),
                                d.get("juego", ""),
                                d.get("id_pesa", ""),
                                d.get("nominal", ""),
                                d.get("unidad", ""),
                            ]
                        )
            else:
                hoja.append(["No se seleccionaron pesas para esta prueba"])

            hoja.append([])

            agregar_titulo_bloque("Medidores")
            hoja.append(
                [
                    "Magnitud",
                    "Valor Inicial (Ci)",
                    "Hora Inicial",
                    "Valor Final (Cf)",
                    "Hora Final",
                ]
            )

            time_stamp_actual = datetime.now().strftime("%H:%M:%S")

            val_presion_interfaz = (
                self.label_presion.cget("text")
                if hasattr(self, "label_presion")
                else "---"
            )
            val_presion_f_interfaz = (
                self.label_presion_f.cget("text")
                if hasattr(self, "label_presion_f")
                else "---"
            )
            val_humedad_interfaz = (
                self.label_higrometro.cget("text")
                if hasattr(self, "label_higrometro")
                else "---"
            )
            val_humedad_f_interfaz = (
                self.label_higrometro_f.cget("text")
                if hasattr(self, "label_higrometro_f")
                else "---"
            )
            val_temp_interfaz = (
                self.label_temperatura.cget("text")
                if hasattr(self, "label_temperatura")
                else "---"
            )
            val_temp_f_interfaz = (
                self.label_temperatura_f.cget("text")
                if hasattr(self, "label_temperatura_f")
                else "---"
            )

            if val_presion_f_interfaz == "---" or val_presion_f_interfaz == "":
                val_presion_f_excel = val_presion_interfaz
                val_humedad_f_excel = val_humedad_interfaz
                val_temp_f_excel = val_temp_interfaz
                hora_inicial_excel = time_stamp_actual
                hora_final_excel = time_stamp_actual
            else:
                val_presion_f_excel = val_presion_f_interfaz
                val_humedad_f_excel = val_humedad_f_interfaz
                val_temp_f_excel = val_temp_f_interfaz
                hora_inicial_excel = getattr(
                    self, "hora_inicio_prueba", time_stamp_actual
                )
                hora_final_excel = time_stamp_actual

            if (
                not hasattr(self, "hora_inicio_prueba")
                or self.hora_inicio_prueba is None
            ):
                self.hora_inicio_prueba = time_stamp_actual

            valores_ambientales = [
                [
                    "Presión",
                    val_presion_interfaz,
                    hora_inicial_excel,
                    val_presion_f_excel,
                    hora_final_excel,
                ],
                [
                    "Humedad",
                    val_humedad_interfaz,
                    hora_inicial_excel,
                    val_humedad_f_excel,
                    hora_final_excel,
                ],
                [
                    "Temperatura",
                    val_temp_interfaz,
                    hora_inicial_excel,
                    val_temp_f_excel,
                    hora_final_excel,
                ],
            ]

            for linea in valores_ambientales:
                hoja.append(linea)

            hoja.append([])

            for col in hoja.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                hoja.column_dimensions[col_letter].width = max(max_len + 4, 15)

            wb.save(ruta_directa)
            CTkMessagebox(
                title="Guardado Exitoso",
                message=f"Se encuentra en:\n'{nombre_hoja}'",
                icon="check",
            )

        except Exception as e:
            CTkMessagebox(
                title="Error de Archivo",
                message=f"No se pudo guardar la información:\n{str(e)}",
                icon="cancel",
            )

    def cargar_datos_desde_excel(self):
        ruta_directa = getattr(self.master, "archivo_calibracion_actual", None)
        if not ruta_directa:
            CTkMessagebox(
                title="Archivo Faltante",
                message="No hay ningún archivo cargado en el sistema.\nPor favor, ve al menú superior primero.",
                icon="warning",
            )
            return

        try:
            wb = openpyxl.load_workbook(ruta_directa, data_only=True)
            hojas_excentricidad = [
                hoja for hoja in wb.sheetnames if hoja.startswith("Excentricidad")
            ]

            if not hojas_excentricidad:
                CTkMessagebox(
                    title="Hoja no encontrada",
                    message="El archivo cargado no contiene ninguna pestaña de 'Excentricidad'.",
                    icon="info",
                )
                return

            hoja_seleccionada = hojas_excentricidad[0]
            if len(hojas_excentricidad) > 1:
                lista_hojas_str = "\n".join([f"- {h}" for h in hojas_excentricidad])
                dialogo = ctk.CTkInputDialog(
                    text=f"Se encontraron las siguientes pruebas:\n{lista_hojas_str}\n\nEscribe el nombre exacto de la prueba que deseas cargar:",
                    title="Seleccionar Prueba",
                )
                seleccion = dialogo.get_input()
                if seleccion and seleccion.strip() in wb.sheetnames:
                    hoja_seleccionada = seleccion.strip()
                elif seleccion is None or seleccion.strip() == "":
                    return
                else:
                    CTkMessagebox(
                        title="Error", message="Nombre inválido.", icon="cancel"
                    )
                    return

            hoja = wb[hoja_seleccionada]

            for i in self.tabla.get_children():
                self.tabla.delete(i)
            for i in self.tabla_promedios.get_children():
                self.tabla_promedios.delete(i)
            for i in self.tabla_sel.get_children():
                self.tabla_sel.delete(i)

            bloque_actual = None
            contador_pesas = 1
            suma_total_pesas = 0.0
            self.llaves_actuales = []

            for fila in hoja.iter_rows(values_only=True):
                if not fila or fila[0] is None:
                    continue
                primer_valor = str(fila[0]).strip()

                if primer_valor == "Indicacion":
                    bloque_actual = "INDICACION"
                    continue
                elif primer_valor == "Promedios por Sección":
                    bloque_actual = "PROMEDIOS"
                    continue
                elif primer_valor == "Pesas":
                    bloque_actual = "PESAS"
                    continue
                elif primer_valor == "Medidores":
                    bloque_actual = "MEDIDORES"
                    continue

                if bloque_actual == "INDICACION":
                    if primer_valor == "Sentido":
                        continue
                    self.tabla.insert(
                        "",
                        "end",
                        values=(
                            str(fila[0]).strip(),
                            str(fila[1]).strip(),
                            fila[2] if fila[2] is not None else "---",
                        ),
                    )

                elif bloque_actual == "PROMEDIOS":
                    if primer_valor == "Sección":
                        continue
                    self.tabla_promedios.insert(
                        "",
                        "end",
                        values=(
                            str(fila[0]).strip(),
                            (
                                f"{fila[1]:.2f}"
                                if isinstance(fila[1], (int, float))
                                else str(fila[1])
                            ),
                        ),
                    )

                elif bloque_actual == "PESAS":
                    if primer_valor in [
                        "Key",
                        "No se seleccionaron pesas para esta prueba",
                    ]:
                        continue

                    id_pesa = str(fila[2]).strip() if fila[2] is not None else ""
                    serie_pesa = str(fila[4]).strip() if fila[4] is not None else ""
                    nominal_pesa = fila[7] if fila[7] is not None else "0"
                    unidad_pesa = str(fila[8]).strip() if fila[8] is not None else "kg"
                    key_pesa = str(fila[0]).strip() if fila[0] is not None else ""

                    if key_pesa:
                        pesa_key_completa = (
                            f"{id_pesa} | {serie_pesa} | {nominal_pesa}{unidad_pesa}"
                        )
                        self.llaves_actuales.append(pesa_key_completa)

                    self.tabla_sel.insert(
                        "",
                        "end",
                        values=(
                            contador_pesas,
                            id_pesa,
                            serie_pesa,
                            nominal_pesa,
                            unidad_pesa,
                        ),
                    )
                    contador_pesas += 1

                    try:
                        val = float(nominal_pesa)
                        if unidad_pesa.lower() in ["g", "gramos", "g."]:
                            suma_total_pesas += val / 1000.0
                        else:
                            suma_total_pesas += val
                    except ValueError:
                        pass

                elif bloque_actual == "MEDIDORES":
                    if primer_valor == "Magnitud":
                        continue
                    val_ci = str(fila[1]).strip() if fila[1] is not None else "---"
                    val_cf = str(fila[2]).strip() if fila[2] is not None else "---"

                    if "Presión" in primer_valor:
                        if hasattr(self, "label_presion"):
                            self.label_presion.configure(text=val_ci)
                        if hasattr(self, "label_presion_f"):
                            self.label_presion_f.configure(text=val_cf)
                    elif "Humedad" in primer_valor:
                        if hasattr(self, "label_higrometro"):
                            self.label_higrometro.configure(text=val_ci)
                        if hasattr(self, "label_higrometro_f"):
                            self.label_higrometro_f.configure(text=val_cf)
                    elif "Temperatura" in primer_valor:
                        if hasattr(self, "label_temperatura"):
                            self.label_temperatura.configure(text=val_ci)
                        if hasattr(self, "label_temperatura_f"):
                            self.label_temperatura_f.configure(text=val_cf)

            if hasattr(self, "label_suma_nominal"):
                self.label_suma_nominal.configure(
                    text=f"Total: {suma_total_pesas:.2f} kg"
                )

            CTkMessagebox(
                title="Éxito",
                message=f"Se cargaron correctamente los datos de '{hoja_seleccionada}'.",
                icon="check",
            )

        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"No se pudieron cargar los datos de excentricidad:\n{str(e)}",
                icon="cancel",
            )

    def nueva_prueba(self):
        msg = CTkMessagebox(
            title="¿Nueva Prueba?",
            message="Se limpiará las tablas de la pantalla actual.",
            icon="question",
            option_1="Sí",
            option_2="No",
        )
        if msg.get() == "Sí":
            for i in self.tabla.get_children():
                self.tabla.delete(i)
            for i in self.tabla_promedios.get_children():
                self.tabla_promedios.delete(i)
            self.sentido_actual = "IDA"
            self.hoja_actual_guardada = None
            if hasattr(self, "entry_secciones"):
                self.entry_secciones.delete(0, "end")
            if hasattr(self, "btn_sentido"):
                self.btn_sentido.configure(text="Sentido: IDA", fg_color="#16a085")
            CTkMessagebox(
                title="Listo",
                message="Pantalla restablecida. Introduce las secciones para iniciar.",
                icon="check",
            )

    def mostrar_selector_pesas(self):
        diccionario_pesas = getattr(self, "pesas", {})

        lista_patrones = [
            llave
            for llave, datos in diccionario_pesas.items()
            if str(datos.get("sustitucion")).strip().lower() == "no"
        ]

        if lista_patrones:
            Ventana_Pesas(self, lista_patrones, self.actualizar_tabla_pesas)
        else:
            CTkMessagebox(
                title="Sin registros",
                message="No se encontraron elementos marcados como patrones del archivo cargado.",
                icon="warning",
            )

    def actualizar_tabla_pesas(self, seleccionadas):
        if not hasattr(self, "llaves_actuales") or self.llaves_actuales is None:
            self.llaves_actuales = []

        for llave in seleccionadas:
            if llave not in self.llaves_actuales:
                self.llaves_actuales.append(llave)

        for item in self.tabla_sel.get_children():
            self.tabla_sel.delete(item)

        suma_total = 0.0
        diccionario_pesas = getattr(self, "pesas", {})

        for i, llave in enumerate(self.llaves_actuales, start=1):
            datos = diccionario_pesas.get(llave)
            if datos:
                self.tabla_sel.insert(
                    "",
                    "end",
                    values=(
                        i,
                        datos["id"],
                        datos["serie"],
                        datos["nominal"],
                        datos["unidad"],
                    ),
                )

                try:
                    val = float(datos["nominal"])
                    unidad = str(datos["unidad"]).lower().strip()
                    if unidad in ["g", "gramos", "g."]:
                        suma_total += val / 1000.0
                    else:
                        suma_total += val
                except ValueError:
                    pass

        self.label_suma_nominal.configure(text=f"Total: {suma_total:.2f} kg")

    def eliminar_ultima_pesa(self):
        if hasattr(self, "llaves_actuales") and self.llaves_actuales:
            self.llaves_actuales.pop()

            for item in self.tabla_sel.get_children():
                self.tabla_sel.delete(item)

            suma_total = 0.0
            diccionario_pesas = getattr(self, "pesas", {})

            for i, llave in enumerate(self.llaves_actuales, start=1):
                datos = diccionario_pesas.get(llave)
                if datos:
                    self.tabla_sel.insert(
                        "",
                        "end",
                        values=(
                            i,
                            datos["id"],
                            datos["serie"],
                            datos["nominal"],
                            datos["unidad"],
                        ),
                    )
                    try:
                        val = float(datos["nominal"])
                        unidad = str(datos["unidad"]).lower().strip()
                        if unidad in ["g", "gramos", "g."]:
                            suma_total += val / 1000.0
                        else:
                            suma_total += val
                    except ValueError:
                        pass

            self.label_suma_nominal.configure(text=f"Total: {suma_total:.2f} kg")

        else:
            CTkMessagebox(
                title="Aviso",
                message="No hay pesas en la lista para eliminar.",
                icon="info",
            )

    def mostrar_selector_sustitucion(self):
        diccionario_pesas = getattr(self, "pesas", {})

        lista_sustitucion = [
            llave
            for llave, datos in diccionario_pesas.items()
            if datos.get("sustitucion") == "si"
        ]

        if lista_sustitucion:
            Ventana_Pesas(self, lista_sustitucion, self.actualizar_tabla_pesas)
        else:
            CTkMessagebox(
                title="Sin registros",
                message="No se encontraron elementos marcados como sustitución archivo cargado.",
                icon="warning",
            )

    def registrar_CI(self):
        T = self.label_temperatura.cget("text")
        H = self.label_higrometro.cget("text")
        P = self.label_presion.cget("text")

        t_inst = self.combo_temperatura.get()
        h_inst = self.combo_higrometro.get()
        p_inst = self.combo_presion.get()

        self.condiciones_iniciales = {
            "Temperatura": {"Valor": T, "Instrumento": t_inst},
            "Humedad": {"Valor": H, "Instrumento": h_inst},
            "Presion": {"Valor": P, "Instrumento": p_inst},
            "Hora": datetime.now().strftime("%H:%M:%S"),
        }
        self.label_barI.configure(text=P, text_color="#158A30")
        self.label_higI.configure(text=H, text_color="#158A30")
        self.label_temI.configure(text=T, text_color="#158A30")

        self.btn_CI.configure(fg_color="#158A30")

    def registrar_CF(self):
        T = self.label_temperatura.cget("text")
        H = self.label_higrometro.cget("text")
        P = self.label_presion.cget("text")

        t_inst = self.combo_temperatura.get()
        h_inst = self.combo_higrometro.get()
        p_inst = self.combo_presion.get()

        self.condiciones_finales = {
            "Temperatura": {"Valor": T, "Instrumento": t_inst},
            "Humedad": {"Valor": H, "Instrumento": h_inst},
            "Presion": {"Valor": P, "Instrumento": p_inst},
            "Hora": datetime.now().strftime("%H:%M:%S"),
        }
        self.label_barF.configure(text=P, text_color="#D95E4A")
        self.label_higF.configure(text=H, text_color="#D95E4A")
        self.label_temF.configure(text=T, text_color="#D95E4A")

        self.btn_CF.configure(fg_color="#D95E4A")

    def eliminar_ultima_indicacion(self):
        filas = self.tabla.get_children()
        if not filas:
            CTkMessagebox(
                title="Aviso",
                message="La tabla está vacía. Genera las secciones primero.",
                icon="info",
            )
            return

        for item in reversed(filas):
            v = self.tabla.item(item)["values"]

            if str(v[2]).strip() != "---":
                self.tabla.item(item, values=(v[0], v[1], "---"))

                self.tabla.selection_set(item)
                self.tabla.see(item)

                self.calcular_promedios_secciones()
                return

        CTkMessagebox(
            title="Aviso",
            message="No hay ninguna indicación registrada para eliminar.",
            icon="info",
        )

    def guardar_como_nuevo_excel(self):

        filas_tabla = self.tabla.get_children()
        if not filas_tabla:
            CTkMessagebox(
                title="Aviso",
                message="No hay datos de indicaciones en la tabla para exportar.",
                icon="warning",
            )
            return

        ruta_guardar = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos de Excel", "*.xlsx")],
            title="Guardar Prueba de Excentricidad",
            initialfile=f"Prueba_Excentricidad_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        )

        if not ruta_guardar:
            return

        try:

            wb = openpyxl.Workbook()
            hoja = wb.active
            hoja.title = "Excentricidad"
            hoja.views.sheetView[0].showGridLines = True

            color_encabezado = PatternFill(
                start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
            )
            fuente_negrita = Font(bold=True)

            def agregar_titulo_bloque(texto):
                hoja.append([texto])
                celda = hoja.cell(row=hoja.max_row, column=1)
                celda.fill = color_encabezado
                celda.font = fuente_negrita

            agregar_titulo_bloque("Indicacion")
            hoja.append(["Sentido", "Sección", "Valor (kg)"])

            for item in filas_tabla:
                v = self.tabla.item(item)["values"]
                val_ind = (
                    float(v[2])
                    if str(v[2]).strip() != "---" and str(v[2]).strip() != ""
                    else v[2]
                )
                hoja.append([v[0], v[1], val_ind])

            hoja.append([])

            agregar_titulo_bloque("Promedios por Sección")
            hoja.append(["Sección", "Promedio (kg)"])

            for item in self.tabla_promedios.get_children():
                vp = self.tabla_promedios.item(item)["values"]
                try:
                    hoja.append([vp[0], float(vp[1])])
                except ValueError:
                    hoja.append([vp[0], vp[1]])

            hoja.append([])

            agregar_titulo_bloque("Pesas")
            hoja.append(
                [
                    "Key",
                    "Magnitud",
                    "Identificación",
                    "Modelo",
                    "Serie",
                    "Juego",
                    "Id_pesa",
                    "Nominal",
                    "Unidad",
                ]
            )

            if hasattr(self, "llaves_actuales") and self.llaves_actuales:
                diccionario_local_pesas = getattr(self, "pesas", {})
                for llave in self.llaves_actuales:
                    d = diccionario_local_pesas.get(llave)
                    if d:
                        hoja.append(
                            [
                                d.get("key", ""),
                                d.get("magnitud", ""),
                                d.get("id", ""),
                                d.get("modelo", ""),
                                d.get("serie", ""),
                                d.get("juego", ""),
                                d.get("id_pesa", ""),
                                d.get("nominal", ""),
                                d.get("unidad", ""),
                            ]
                        )
            else:
                hoja.append(["No se seleccionaron pesas para esta prueba"])

            hoja.append([])

            agregar_titulo_bloque("Medidores")
            hoja.append(
                [
                    "Magnitud",
                    "Valor Inicial (Ci)",
                    "Hora Inicial",
                    "Valor Final (Cf)",
                    "Hora Final",
                ]
            )

            time_stamp_actual = datetime.now().strftime("%H:%M:%S")

            val_presion_interfaz = (
                self.label_presion.cget("text")
                if hasattr(self, "label_presion")
                else "---"
            )
            val_presion_f_interfaz = (
                self.label_presion_f.cget("text")
                if hasattr(self, "label_presion_f")
                else "---"
            )
            val_humedad_interfaz = (
                self.label_higrometro.cget("text")
                if hasattr(self, "label_higrometro")
                else "---"
            )
            val_humedad_f_interfaz = (
                self.label_higrometro_f.cget("text")
                if hasattr(self, "label_higrometro_f")
                else "---"
            )
            val_temp_interfaz = (
                self.label_temperatura.cget("text")
                if hasattr(self, "label_temperatura")
                else "---"
            )
            val_temp_f_interfaz = (
                self.label_temperatura_f.cget("text")
                if hasattr(self, "label_temperatura_f")
                else "---"
            )

            if val_presion_f_interfaz == "---" or val_presion_f_interfaz == "":
                val_presion_f_excel = val_presion_interfaz
                val_humedad_f_excel = val_humedad_interfaz
                val_temp_f_excel = val_temp_interfaz
                hora_inicial_excel = time_stamp_actual
                hora_final_excel = time_stamp_actual
            else:
                val_presion_f_excel = val_presion_f_interfaz
                val_humedad_f_excel = val_humedad_f_interfaz
                val_temp_f_excel = val_temp_f_interfaz
                hora_inicial_excel = getattr(
                    self, "hora_inicio_prueba", time_stamp_actual
                )
                hora_final_excel = time_stamp_actual

            valores_ambientales = [
                [
                    "Presión",
                    val_presion_interfaz,
                    hora_inicial_excel,
                    val_presion_f_excel,
                    hora_final_excel,
                ],
                [
                    "Humedad",
                    val_humedad_interfaz,
                    hora_inicial_excel,
                    val_humedad_f_excel,
                    hora_final_excel,
                ],
                [
                    "Temperatura",
                    val_temp_interfaz,
                    hora_inicial_excel,
                    val_temp_f_excel,
                    hora_final_excel,
                ],
            ]

            for linea in valores_ambientales:
                hoja.append(linea)

            for col in hoja.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                hoja.column_dimensions[col_letter].width = max(max_len + 4, 15)

            wb.save(ruta_guardar)

            self.master.archivo_calibracion_actual = ruta_guardar
            self.hoja_actual_guardada = "Excentricidad"

            CTkMessagebox(
                title="Éxito",
                message=f"Archivo nuevo creado y guardado con éxito en:\n{ruta_guardar}",
                icon="check",
            )

        except Exception as e:
            CTkMessagebox(
                title="Error al exportar",
                message=f"No se pudo crear el archivo nuevo:\n{str(e)}",
                icon="cancel",
            )


class Error_Indicacion(ctk.CTkToplevel):
    """
    Ventana para el error de indicacion
    """

    def __init__(self, parent, master, shell, lista_instrumentos=None, ruta_excel=None):
        super().__init__(parent)
        self.master = master
        self.shell = shell
        self.lista_instrumentos = lista_instrumentos

        if ruta_excel:
            setattr(self, "ruta_excel", ruta_excel)

        self.llaves_actuales = []
        self.pesas_del_step_actual = 0.0
        self.hoja_actual_guardada = None
        self.sustitucion = False
        self.hora_inicio = None

        self.pesas = {}
        self.lista_pesas = []

        if ruta_excel:
            self.cargar_modelos(ruta_excel)
        self.title("Error de Indicación")
        self.geometry("1390x800+70+20")
        self.grab_set()
        self.menu()
        self.widgets()
        self.actualizar_peso()

    def cargar_modelos(self, ruta_excel):
        try:
            wb = openpyxl.load_workbook(ruta_excel, data_only=True)
            if "Equipo Pesas" in wb.sheetnames:
                hoja_pesas = wb["Equipo Pesas"]
                for i in hoja_pesas.iter_rows(min_row=2, max_col=12, values_only=True):
                    if i[2] is not None:
                        c = str(i[2]).strip()
                        f = str(i[5]).strip() if i[5] is not None else ""
                        j = str(i[9]).strip() if i[9] is not None else ""
                        k = str(i[10]).strip() if i[10] is not None else ""

                        sust = str(i[11]).strip().lower() if i[11] is not None else "no"
                        pesa_key = f"{c} | {f} | {j}{k}"

                        self.pesas[pesa_key] = {
                            "key": str(i[0]).strip(),
                            "magnitud": str(i[1]).strip() if i[1] else "",
                            "id": c,
                            "modelo": str(i[4]).strip() if i[4] else "",
                            "serie": f,
                            "juego": str(i[6]).strip() if i[6] else "",
                            "id_pesa": str(i[7]).strip() if i[7] else "",
                            "nominal": j,
                            "unidad": k,
                            "sustitucion": sust,
                        }
                        self.lista_pesas_completa.append(pesa_key)
        except Exception:
            pass

    def widgets(self):
        self.frame_repetibilidad = ctk.CTkFrame(self, width=800, height=700)
        self.frame_repetibilidad.pack(fill="both", expand=True, padx=5, pady=5)
        self.frame_repetibilidad.pack_propagate(False)

        self.frame_peso = ctk.CTkFrame(self.frame_repetibilidad, width=400, height=150)
        self.frame_peso.pack(anchor="n", padx=10, pady=10, fill="x")
        self.frame_peso.pack_propagate(False)

        self.frame_peso_instantaneo = ctk.CTkFrame(
            self.frame_peso, width=450, height=150
        )
        self.frame_peso_instantaneo.pack(side="left", padx=10, pady=10)
        self.frame_peso_instantaneo.pack_propagate(False)

        self.label_unidad = ctk.CTkLabel(
            self.frame_peso_instantaneo, text="kg", font=("Cambria", 50)
        )
        self.label_unidad.pack(side="right", padx=20, pady=5)

        self.frame_peso_al_momento = ctk.CTkFrame(
            self.frame_peso_instantaneo, width=450, height=150
        )
        self.frame_peso_al_momento.pack(side="left", padx=10, pady=10)
        self.frame_peso_al_momento.pack_propagate(False)

        self.label_peso = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 70)
        )
        self.label_peso.pack(padx=20, pady=5, expand=True)

        self.label_tara = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 12)
        )
        self.label_tara.place(x=8, y=8)

        self.label_neto = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 12)
        )
        self.label_neto.place(x=8, y=79)

        # ---------BOTONES-------------------
        imagen_tarar = Image.open(ruta(os.path.join("Icon", "aumentar.png")))
        icon_tarar = ctk.CTkImage(
            light_image=imagen_tarar, dark_image=imagen_tarar, size=(30, 30)
        )

        self.button_tarar = ctk.CTkButton(
            self.frame_peso,
            text="Tara",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.tara,
            image=icon_tarar,
            compound="top",
        )
        self.button_tarar.pack(side="left", expand=True)
        # ---------------BOTON ZERO-------------
        imagen_zero = Image.open(ruta(os.path.join("Icon", "cero.png")))
        icon_zero = ctk.CTkImage(
            light_image=imagen_zero, dark_image=imagen_zero, size=(30, 30)
        )

        self.button_zero = ctk.CTkButton(
            self.frame_peso,
            text="Zero",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.zero,
            image=icon_zero,
            compound="top",
        )
        self.button_zero.pack(side="left", expand=True)

        # ---------------BOTON QUITAR TARAR-------------
        imagen_q_tara = Image.open(ruta(os.path.join("Icon", "perdida-peso.png")))
        icon_q_tara = ctk.CTkImage(
            light_image=imagen_q_tara, dark_image=imagen_q_tara, size=(30, 30)
        )
        self.button_quitar_tara = ctk.CTkButton(
            self.frame_peso,
            text="Eliminar Tara",
            width=2,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.quitar_tara,
            image=icon_q_tara,
            compound="top",
        )
        self.button_quitar_tara.pack(side="left", expand=True)

        # ----------------Condiciones Ambientales-----------------
        self.frame_condiciones = ctk.CTkFrame(
            self.frame_repetibilidad, width=400, height=150
        )
        self.frame_condiciones.pack(anchor="s", padx=10, pady=5, fill="x")
        self.frame_condiciones.pack_propagate(False)

        self.frame_condicion1 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion1.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion1.pack_propagate(False)

        self.label_condiciones = ctk.CTkLabel(
            self.frame_condicion1,
            text="Presión",
            font=("Cambria", 20),
        )
        self.label_condiciones.pack(anchor="n", padx=2, pady=5, expand=True)

        self.frame_condicion2 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion2.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion2.pack_propagate(False)

        self.label_condiciones2 = ctk.CTkLabel(
            self.frame_condicion2,
            text="Humedad",
            font=("Cambria", 20),
        )
        self.label_condiciones2.pack(anchor="n", padx=2, pady=5, expand=True)

        self.frame_condicion3 = ctk.CTkFrame(
            self.frame_condiciones, width=220, height=150
        )
        self.frame_condicion3.pack(side="left", padx=10, pady=10, expand=True)
        self.frame_condicion3.pack_propagate(False)

        self.label_condiciones3 = ctk.CTkLabel(
            self.frame_condicion3,
            text="Temperatura",
            font=("Cambria", 20),
        )
        self.label_condiciones3.pack(anchor="n", padx=2, pady=5, expand=True)
        # ---------PRESION-----------------

        self.frame_presion = ctk.CTkFrame(self.frame_condicion1, width=200, height=50)
        self.frame_presion.pack(expand=True, anchor="n")
        self.frame_presion.pack_propagate(False)

        self.label_presion = ctk.CTkLabel(
            self.frame_presion,
            text="---",
            font=("Cambria", 30),
        )
        self.label_presion.pack(expand=True)

        self.combo_presion = ctk.CTkOptionMenu(
            master=self.frame_condicion1,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_presion.pack(expand=True)
        self.combo_presion.set("Modelos")

        # -------------HIGROMETRO-----------------

        self.frame_higrometro = ctk.CTkFrame(
            self.frame_condicion2, width=200, height=50
        )
        self.frame_higrometro.pack(expand=True, anchor="n")
        self.frame_higrometro.pack_propagate(False)

        self.label_higrometro = ctk.CTkLabel(
            self.frame_higrometro,
            text="---",
            font=("Cambria", 30),
        )
        self.label_higrometro.pack(expand=True)
        self.combo_higrometro = ctk.CTkOptionMenu(
            master=self.frame_condicion2,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_higrometro.pack(expand=True)
        self.combo_higrometro.set("Modelos")

        # --------------Temperatura-----------------

        self.frame_temperatura = ctk.CTkFrame(
            self.frame_condicion3, width=200, height=50
        )
        self.frame_temperatura.pack(expand=True, anchor="n")
        self.frame_temperatura.pack_propagate(False)

        self.label_temperatura = ctk.CTkLabel(
            self.frame_temperatura, text="---", font=("Cambria", 30)
        )
        self.label_temperatura.pack(expand=True)
        self.combo_temperatura = ctk.CTkOptionMenu(
            master=self.frame_condicion3,
            values=self.lista_instrumentos,
            width=120,
            height=25,
            font=("Cambria", 12),
        )
        self.combo_temperatura.pack(expand=True)
        self.combo_temperatura.set("Modelos")

        # --------------------------------
        self.frame_registro = ctk.CTkFrame(
            self.frame_repetibilidad, width=400, height=420
        )
        self.frame_registro.pack(anchor="s", padx=10, pady=5, fill="x")
        self.frame_registro.pack_propagate(False)
        # ------tabla registros-----------------
        self.frame_tablas = ctk.CTkFrame(self.frame_registro, fg_color="transparent")
        self.frame_tablas.pack(side="left", fill="both", expand=True, padx=5)

        self.style_tabla = ttk.Style()
        self.style_tabla.theme_use("clam")
        self.style_tabla.configure(
            "Treeview",
            background="#BD9D9D",
            foreground="black",
            rowheight=25,
            fieldbackground="#DFF2F5",
            font=("Cambria", 10),
        )
        self.style_tabla.map(
            "Treeview",
            background=[("selected", "#158A30")],
            foreground=[("selected", "white")],
        )

        columnas_patrones = ("#", "LTj", "Ij", "Ej", "I_prima_j")
        self.tabla_patrones = ttk.Treeview(
            self.frame_tablas, columns=columnas_patrones, show="headings", height=12
        )

        self.tabla_patrones.heading("#", text="#")
        self.tabla_patrones.heading("LTj", text="Lₜⱼ (kg)")
        self.tabla_patrones.heading("Ij", text="Iⱼ (kg)")
        self.tabla_patrones.heading("Ej", text="Eⱼ (kg)")
        self.tabla_patrones.heading("I_prima_j", text="I'ⱼ (kg)")

        for col in columnas_patrones:
            self.tabla_patrones.column(col, width=75, anchor="center")
        self.tabla_patrones.column("#", width=45)
        self.tabla_patrones.pack(side="left", padx=(5, 0), pady=10, fill="y")

        columnas_sustitucion = ("I_Lsubj", "delta_Ij", "Lsubj")
        self.tabla_sustitucion = ttk.Treeview(
            self.frame_tablas, columns=columnas_sustitucion, show="headings", height=12
        )

        self.tabla_sustitucion.heading("I_Lsubj", text="I(Lₛᵤ₆ₜⱼ)")
        self.tabla_sustitucion.heading("delta_Ij", text="ΔIⱼ (kg)")
        self.tabla_sustitucion.heading("Lsubj", text="Lₛᵤ₆ₜⱼ (kg)")

        for col in columnas_sustitucion:
            self.tabla_sustitucion.column(col, width=82, anchor="center")
        self.tabla_sustitucion.pack(side="left", padx=(0, 5), pady=10, fill="y")

        # -----#-----#_-------
        # columnas = ("#", "LTj", "Ij", "Ej", "I_prima_j", "I_Lsubj", "delta_Ij", "Lsubj")

        # self.tabla = ttk.Treeview(
        #     self.frame_registro,
        #     columns=columnas,
        #     show="headings",
        #     height=10,
        # )
        # self.tabla.heading("#", text="#")
        # self.tabla.heading("LTj", text="L\u209c\u1d0a")
        # self.tabla.heading("Ij", text="I\u1d0a")
        # self.tabla.heading("Ej", text="E\u1d0a")
        # self.tabla.heading("I_prima_j", text="I'\u1d0a")
        # self.tabla.heading("I_Lsubj", text="I(L\u209b\u1d6a\u209c\u1d0a)")
        # self.tabla.heading("delta_Ij", text="\u0394I\u1d0a")
        # self.tabla.heading("Lsubj", text="L\u209b\u1d6a\u209c\u1d0a")

        # self.tabla.column("#", width=40, anchor="center")
        # self.tabla.column("LTj", width=80, anchor="center")
        # self.tabla.column("Ij", width=80, anchor="center")
        # self.tabla.column("Ej", width=80, anchor="center")
        # self.tabla.column("I_prima_j", width=80, anchor="center")
        # self.tabla.column("I_Lsubj", width=80, anchor="center")
        # self.tabla.column("delta_Ij", width=80, anchor="center")
        # self.tabla.column("Lsubj", width=80, anchor="center")
        # self.tabla.pack(side="left", padx=10, pady=20, anchor="n")
        # self.tabla = ttk.Treeview(
        #     self.frame_registro,
        #     columns=("#", "Ltj", "Ij", "I(L)"),
        #     show="headings",
        #     height=10,
        # )

        # self.tabla.heading("#", text="#")
        # self.tabla.heading("Ltj", text="L\u209c\u209a")
        # self.tabla.heading("Ij", text="I\u1d0a")
        # self.tabla.heading("I(L)", text="I(L)")

        # self.tabla.column("#", width=40, anchor="center")
        # self.tabla.column("Ltj", width=120, anchor="center")
        # self.tabla.column("Ij", width=120, anchor="center")
        # self.tabla.column("I(L)", width=120, anchor="center")
        # self.tabla.pack(side="left", padx=10, pady=20, anchor="n")

        # -------BOTONES INDICACIONES-----------------

        self.frame_botones_EI = ctk.CTkFrame(self.frame_registro, width=390, height=130)
        self.frame_botones_EI.place(x=620, y=5)
        # self.frame_botones_EI.pack(side="left", padx=5, pady=20, anchor="n")
        self.frame_botones_EI.pack_propagate(False)

        imagen_borrar = Image.open(ruta(os.path.join("Icon", "borrarRojo.png")))
        icon_borrar = ctk.CTkImage(
            light_image=imagen_borrar, dark_image=imagen_borrar, size=(30, 30)
        )

        self.borrar_ultimo = ctk.CTkButton(
            self.frame_botones_EI,
            text="Eliminar punto",
            command=self.borrar_ultima_indicacion,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_borrar,
        )

        self.borrar_ultimo.grid(row=0, column=0, padx=5, pady=10)

        imagen_generar = Image.open(ruta(os.path.join("Icon", "idea.png")))
        icon_generar = ctk.CTkImage(
            light_image=imagen_generar, dark_image=imagen_generar, size=(30, 30)
        )

        self.btn_generar = ctk.CTkButton(
            self.frame_botones_EI,
            text="Generar",
            command=self.añadir_nuevo_step_dinamico,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_generar,
        )
        self.btn_generar.grid(row=1, column=0, padx=5, pady=10)

        imagen_registro = Image.open(ruta(os.path.join("Icon", "boton-agregar.png")))
        icon_registro = ctk.CTkImage(
            light_image=imagen_registro, dark_image=imagen_registro, size=(30, 30)
        )

        self.btn_registro = ctk.CTkButton(
            self.frame_botones_EI,
            text="Registrar peso",
            command=self.registrar_peso_estable,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_registro,
        )
        self.btn_registro.grid(row=2, column=0, padx=5, pady=10)

        # --------BOTONES DE PESAS SUSTITUCION O NADA----

        imagen_pesas = Image.open(ruta(os.path.join("Icon", "escoger.png")))
        icon_pesas = ctk.CTkImage(
            light_image=imagen_pesas, dark_image=imagen_pesas, size=(30, 30)
        )

        self.btn_pesas = ctk.CTkButton(
            self.frame_botones_EI,
            text="Pesas",
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.mostrar_selector_pesas,
            image=icon_pesas,
        )
        self.btn_pesas.grid(row=0, column=1, padx=5, pady=10)

        self.btn_sustitucion = ctk.CTkButton(
            self.frame_botones_EI,
            text="Sustitución",
            command=self.mostrar_selector_sustitucion,
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            image=icon_pesas,
        )
        self.btn_sustitucion.grid(row=1, column=1, padx=5, pady=10)

        self.btn_registro_sustitucion = ctk.CTkButton(
            self.frame_botones_EI,
            text="Susticion Manual",
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            fg_color="#8E44AD",
            hover_color="#7D3C98",
            command=self.registrar_sustitucion,
        )

        self.btn_registro_sustitucion.grid(row=2, column=1, padx=5, pady=10)

        imagen_borrar_pesa = Image.open(ruta(os.path.join("Icon", "borrar.png")))
        icon_borrar_pesas = ctk.CTkImage(
            light_image=imagen_borrar_pesa, dark_image=imagen_borrar_pesa, size=(30, 30)
        )

        self.btn_borrar_pesa = ctk.CTkButton(
            self.frame_botones_EI,
            text="Borrar Pesas",
            width=150,
            height=10,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.borrar_pesas,
            image=icon_borrar_pesas,
        )

        self.btn_borrar_pesa.grid(row=3, column=0, padx=5, pady=10)

        # ----------Tabla de pesos -----------

        self.tabla_sel = ttk.Treeview(
            self.frame_registro,
            columns=("#", "ID", "serie", "valor", "unidad"),
            show="headings",
            height=8,
        )
        self.tabla_sel.heading("#", text="#")
        self.tabla_sel.heading("ID", text="Identificación")
        self.tabla_sel.heading("serie", text="Serie")
        self.tabla_sel.heading("valor", text="V nominal")
        self.tabla_sel.heading("unidad", text="Unidad")

        self.tabla_sel.column("#", width=30, anchor="center")
        self.tabla_sel.column("ID", width=90, anchor="center")
        self.tabla_sel.column("serie", width=90, anchor="center")
        self.tabla_sel.column("valor", width=90, anchor="center")
        self.tabla_sel.column("unidad", width=70, anchor="center")
        self.tabla_sel.place(x=950, y=10)
        # self.tabla_sel.pack(side="left", padx=5, pady=20, anchor="n")

        self.label_suma_nominal = ctk.CTkLabel(
            self.frame_registro, text="Suma Nominal: 0 kg", font=("Cambria", 12)
        )
        self.label_suma_nominal.place(x=1160, y=250)

        # -----------------------------------------

        # _------------------condiciones----
        self.frame_R_condiciones = ctk.CTkFrame(
            self.frame_registro, width=450, height=45
        )
        self.frame_R_condiciones.place(x=800, y=290)
        self.frame_R_condiciones.pack_propagate(False)

        self.frame_r_bar = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_bar.pack(side="left", expand=True, padx=10)
        self.frame_r_bar.pack_propagate(False)

        self.label_barI = ctk.CTkLabel(
            self.frame_r_bar, text="--", font=("Cambria", 15)
        )
        self.label_barI.pack(expand=True)

        self.frame_r_pres = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_pres.pack(side="left", expand=True)
        self.frame_r_pres.pack_propagate(False)

        self.label_higI = ctk.CTkLabel(
            self.frame_r_pres, text="--", font=("Cambria", 15)
        )
        self.label_higI.pack(expand=True)

        self.frame_r_tem = ctk.CTkFrame(self.frame_R_condiciones, width=100, height=30)
        self.frame_r_tem.pack(side="left", expand=True)
        self.frame_r_tem.pack_propagate(False)

        self.label_temI = ctk.CTkLabel(
            self.frame_r_tem, text="--", font=("Cambria", 15)
        )
        self.label_temI.pack(expand=True)

        # ---------condiciones finales ------

        self.frame_cf = ctk.CTkFrame(self.frame_registro, width=450, height=45)
        self.frame_cf.place(x=800, y=350)
        self.frame_cf.pack_propagate(False)

        self.frame_barF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_barF.pack(side="left", expand=True, padx=10)
        self.frame_barF.pack_propagate(False)

        self.label_barF = ctk.CTkLabel(
            self.frame_barF, text="---", font=("Cambria", 15)
        )
        self.label_barF.pack(expand=True)

        self.frame_higF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_higF.pack(side="left", expand=True)
        self.frame_higF.pack_propagate(False)
        self.label_higF = ctk.CTkLabel(
            self.frame_higF, text="---", font=("Cambria", 15)
        )
        self.label_higF.pack(expand=True)

        self.frame_temF = ctk.CTkFrame(self.frame_cf, width=100, height=30)
        self.frame_temF.pack(side="left", expand=True)
        self.frame_temF.pack_propagate(False)

        self.label_temF = ctk.CTkLabel(
            self.frame_temF, text="---", font=("Cambria", 15)
        )
        self.label_temF.pack(expand=True)
        # -------BOTONES CI/cf
        imagen_C = Image.open(ruta(os.path.join("Icon", "escribir.png")))
        icon_C = ctk.CTkImage(light_image=imagen_C, dark_image=imagen_C, size=(30, 30))

        self.btn_CI = ctk.CTkButton(
            self.frame_botones_EI,
            text="Registrar CI",
            width=150,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            command=self.registrar_CI,
            image=icon_C,
        )
        self.btn_CI.grid(row=3, column=1)

        self.btn_CF = ctk.CTkButton(
            self.frame_botones_EI,
            text="Registrar CF",
            width=150,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            command=self.registrar_CF,
            image=icon_C,
        )

        self.btn_CF.grid(row=4, column=1, padx=5, pady=10)

        imagen_guardar = Image.open(
            ruta(os.path.join("Icon", "guardar-el-archivo.png"))
        )
        icon_guardar = ctk.CTkImage(
            light_image=imagen_guardar, dark_image=imagen_guardar, size=(30, 30)
        )

        self.btn_exportar_excel = ctk.CTkButton(
            self.frame_botones_EI,
            text="Guardar",
            width=150,
            height=10,
            corner_radius=20,
            border_color="white",
            border_width=1,
            fg_color="#1F4E79",
            hover_color="#F75036",
            command=self.guardar_xlsx,
            image=icon_guardar,
        )
        self.btn_exportar_excel.grid(row=3, column=0, padx=5, pady=10)

    def menu(self):
        self.menu = CTkMenuBar(master=self)
        self.b1 = self.menu.add_cascade("Archivo", command=self.abrir_modelos)
        self.b2_puerto_bar = self.menu.add_cascade("Barometro")
        self.dropdown_puerto_bar = CustomDropdownMenu(widget=self.b2_puerto_bar)
        self.b3_puerto_higro = self.menu.add_cascade("Higrometro")
        self.dropdown_puerto_higro = CustomDropdownMenu(widget=self.b3_puerto_higro)
        self.abrir = self.menu.add_cascade("Abrir archivo", command=self.abrir_xlsx)
        self.b_nueva_prueba = self.menu.add_cascade(
            "Nueva prueba", command=self.nueva_prueba
        )
        self.b4 = self.menu.add_cascade(
            "Leer xlsx", command=self.master.cargar_excel_calibracion
        )
        self.b5 = self.menu.add_cascade(
            "Cargar Prueba", command=self.cargar_datos_desde_excel
        )
        # self.b6 = self.menu.add_cascade("Nueva Prueba", command=self.nueva_prueba)

        puertos = serial.tools.list_ports.comports()
        if not puertos:
            self.dropdown_puerto_bar.add_command(label="No se encontraron puertos")
            self.dropdown_puerto_higro.add_command(label="No se encontraron puertos")
        else:
            for i in puertos:
                self.dropdown_puerto_bar.add_option(
                    option=i.device,
                    command=lambda puerto=i.device: self.seleccionar_puerto_bar(puerto),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                )
                self.dropdown_puerto_higro.add_option(
                    option=i.device,
                    command=lambda puerto=i.device: self.seleccionar_puerto_higro(
                        puerto
                    ),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                )

    def actualizar_peso(self):
        global Lectura_barometro, LecturaT1, LecturaT2
        try:
            val_menos = "{:.2f}".format(float(Lectura_barometro))
            self.label_presion.configure(text=f"{val_menos} Pa")
            self.label_temperatura.configure(text=f"{LecturaT1} °C")
            self.label_higrometro.configure(text=f"{LecturaH1} %")
        except Exception as e:
            print(f"error -{e}")

        lecture = self.shell.peso_instantaneo()

        if len(lecture) >= 3:
            try:
                peso_r = int(float(lecture[2]))
                self.label_peso.configure(text=str(peso_r))
                self.label_neto.configure(text=f"Neto: {peso_r} kg", text_color="white")
            except ValueError:
                pass

        if len(lecture) >= 3:
            wg = lecture[2]
            self.label_peso.configure(text=wg)

            if lecture[1] == "S":
                self.label_peso.configure(text_color="#158A30")
            else:
                self.label_peso.configure(text_color="#B52C19")

            valor_tara = self.shell.obtener_tara()
            self.label_tara.configure(text=f"Tara: {valor_tara}", text_color="#FF6021")

        self.after(500, self.actualizar_peso)

    def abrir_modelos(self):
        ruta = askopenfilename(
            title="Seleccionar archivo", filetypes=[("Archivos de Excel", "*.xlsx")]
        )

        if ruta:
            try:
                wb = openpyxl.load_workbook(ruta, data_only=True)

                if "Equipo CA" in wb.sheetnames:
                    hoja = wb["Equipo CA"]
                    self.equipos_ca = {}
                else:
                    CTkMessagebox(
                        title="Error",
                        message='La hoja "Equipo CA" no se encuentra en el archivo',
                        icon="error",
                    )
                    return

                lista_sensores = []
                for fila in hoja.iter_rows(min_row=2, max_col=10, values_only=True):
                    if fila[0] is not None:
                        a = str(fila[0]).strip()
                        b = str(fila[1]).strip() if fila[1] is not None else ""
                        c = str(fila[2]).strip() if fila[2] is not None else ""
                        d = str(fila[3]).strip() if fila[3] is not None else ""
                        e = str(fila[5]).strip() if fila[5] is not None else ""
                        h = str(fila[7]).strip() if fila[7] is not None else ""
                        sensor_info = f"{d}, {e}, {h}".strip(", ")

                        self.equipos_ca[sensor_info] = {
                            "key": a,
                            "magnitud": b,
                            "indicacion": c,
                            "marca": d,
                            "modelo": e,
                            "serie_sensor": h,
                        }

                        if sensor_info not in lista_sensores:
                            lista_sensores.append(sensor_info)
                self.lista_medidores = lista_sensores
                self.combo_presion.configure(values=lista_sensores)
                self.combo_higrometro.configure(values=lista_sensores)
                self.combo_temperatura.configure(values=lista_sensores)

                if "Equipo Pesas" in wb.sheetnames:
                    hoja_pesas = wb["Equipo Pesas"]
                    self.pesas = {}  # Diccionario maestro local
                    self.lista_pesas_completa = []

                    for idx, i in enumerate(
                        hoja_pesas.iter_rows(min_row=2, max_col=12, values_only=True)
                    ):
                        if i[2] is not None:
                            a = str(i[0]).strip()
                            b = str(i[1]).strip() if i[1] is not None else ""
                            c = str(i[2]).strip() if i[2] is not None else ""
                            e = str(i[4]).strip() if i[4] is not None else ""
                            f = str(i[5]).strip() if i[5] is not None else ""
                            g = str(i[6]).strip() if i[6] is not None else ""
                            h = str(i[7]).strip() if i[7] is not None else ""
                            j = str(i[9]).strip() if i[9] is not None else ""
                            k = str(i[10]).strip() if i[10] is not None else ""

                            sust = (
                                str(i[11]).strip().lower()
                                if i[11] is not None
                                else "no"
                            )

                            pesa_key = f"{c} | {f} | {j}{k}"
                            self.pesas[pesa_key] = {
                                "key": a,
                                "magnitud": b,
                                "id": c,
                                "modelo": e,
                                "serie": f,
                                "juego": g,
                                "id_pesa": h,
                                "nominal": j,
                                "unidad": k,
                                "sustitucion": sust,
                            }
                            self.lista_pesas_completa.append(pesa_key)

                CTkMessagebox(
                    title="Exito",
                    message="Modelos cargados correctamente.",
                    icon="check",
                )

            except Exception as e:
                CTkMessagebox(
                    title="Error",
                    message=f"Error al abrir el archivo: {e}",
                    icon="cancel",
                )
                return

    def seleccionar_puerto_higro(self, puerto):
        if conectar_higrometro(puerto):
            self.hilo_higro = threading.Thread(target=read_hidro, daemon=True)
            self.hilo_higro.start()
            CTkMessagebox(
                title="Higrometro", message=f"conexion con {puerto}", icon="check"
            )
        else:
            print("mal higro")

    def seleccionar_puerto_bar(self, puerto):

        if conectar_barometro(puerto):
            self.hilo_bar = threading.Thread(target=read_bar, daemon=True)
            self.hilo_bar.start()
            CTkMessagebox(
                title="Barometro",
                message=f"Barometro conectado a {puerto}",
                icon="check",
            )
        else:
            CTkMessagebox(title="Barometro", message="Sin conexion", icon="cancel")

    def agregar_punto(self):

        if not self.tabla.get_children():
            self.tabla.insert(
                "", "end", values=("0", "0", "0", "0", "0", "0", "0", "0")
            )
            self.tabla.selection_set(self.tabla.get_children()[0])
            return

        fin = self.tabla.item(self.tabla.get_children()[-1])["values"]
        siguiente = int(fin[0]) + 1

        if str(fin[5]).strip() == "---" and str(fin[0]).strip() != "0":
            CTkMessagebox(
                title="Error",
                message="Completa todas las fases del paso actual antes de agregar un nuevo punto.",
                icon="cancel",
            )
            return

        n = self.tabla.insert(
            "",
            "end",
            values=(str(siguiente), "---", "---", "---", "---", "---", "---", "---"),
        )
        self.tabla.selection_set(n)
        self.tabla.see(n)

        # def registrar_peso_estable(self):
        #     lectura = self.shell.peso_estable()
        #     if len(lectura) < 3:
        #         CTkMessagebox(
        #             title="Error", message="Peso inestable", icon="cancel"
        #         )
        #         return

        #     peso = lectura[2]

        #     for i in self.tabla.get_children():
        #         v = self.tabla.item(i)["values"]
        #         if str(v[0])=="0":continue

        CTkMessagebox(
            title="Aviso",
            message="Todas las fases de la prueba por sustitución están completas.",
            icon="info",
        )

    def registrar_peso_estable(self):
        lectura = self.shell.peso_estable()
        if len(lectura) < 3:
            CTkMessagebox(
                title="Error", message="Peso inestable en la plataforma.", icon="cancel"
            )
            return

        peso_bascula = float(lectura[2])

        seleccion_actual = self.tabla_patrones.selection()
        if not seleccion_actual:
            CTkMessagebox(
                title="Aviso",
                message="Por favor, selecciona el Step activo en la tabla.",
                icon="info",
            )
            return

        item_pat_activo = seleccion_actual[0]
        filas_pat = list(self.tabla_patrones.get_children())
        filas_sust = list(self.tabla_sustitucion.get_children())

        idx = filas_pat.index(item_pat_activo)
        item_sust_activo = filas_sust[idx]

        v_pat = self.tabla_patrones.item(item_pat_activo)["values"]
        v_sust = self.tabla_sustitucion.item(item_sust_activo)["values"]

        if str(v_pat[0]) == "0":
            return

        item_sust_anterior = filas_sust[idx - 1]
        v_sust_anterior = self.tabla_sustitucion.item(item_sust_anterior)["values"]

        if str(v_sust_anterior[2]).strip() != "---":
            l_sub_anterior = float(v_sust_anterior[2])
            es_flujo_sustitucion_activo = True
        else:
            l_sub_anterior = 0.0
            es_flujo_sustitucion_activo = False

        total_patrones_actuales = 0.0
        tiene_sustitucion_si = False

        if hasattr(self, "llaves_actuales") and self.llaves_actuales:
            for llave in self.llaves_actuales:
                datos_p = self.pesas.get(llave, {})
                try:
                    v_nom = float(datos_p.get("nominal", 0))
                    if str(datos_p.get("unidad")).lower().strip() in ["g", "g."]:
                        v_nom = v_nom / 1000.0

                    if str(datos_p.get("sustitucion")).strip().lower() == "si":
                        tiene_sustitucion_si = True
                    else:
                        total_patrones_actuales += v_nom
                except ValueError:
                    pass

        es_metodo_sustitucion_aqui = (
            es_flujo_sustitucion_activo
            or tiene_sustitucion_si
            or getattr(self, "forzar_modo_sustitucion", False)
        )

        if str(v_pat[2]).strip() == "---":
            if es_flujo_sustitucion_activo:
                l_j_total = l_sub_anterior + total_patrones_actuales
            else:
                l_j_total = total_patrones_actuales

            e_j = peso_bascula - l_j_total

            self.tabla_patrones.item(
                item_pat_activo,
                values=(
                    v_pat[0],
                    f"{l_j_total:.2f}",
                    f"{peso_bascula:.2f}",
                    f"{e_j:.2f}",
                    "---",
                ),
            )
            self.tabla_sustitucion.item(item_sust_activo, values=("---", "---", "---"))
            return

        if str(v_pat[4]).strip() == "---":
            self.tabla_patrones.item(
                item_pat_activo,
                values=(v_pat[0], v_pat[1], v_pat[2], v_pat[3], f"{peso_bascula:.2f}"),
            )

            if not es_metodo_sustitucion_aqui:
                if idx + 1 < len(filas_pat):
                    self.tabla_patrones.selection_set(filas_pat[idx + 1])
                    self.tabla_sustitucion.selection_set(filas_sust[idx + 1])
                    self.tabla_patrones.see(filas_pat[idx + 1])
            return

        if str(v_sust[0]).strip() == "---":
            if not es_metodo_sustitucion_aqui:
                CTkMessagebox(
                    title="Método Directo",
                    message="Este paso ya cuenta con Carga y Descarga. Si requieres Sustitución, activa el Switch lateral.",
                    icon="info",
                )
                return

            l_tj_fijado = float(v_pat[1])
            i_j = float(v_pat[2])

            delta_i_j = peso_bascula - i_j
            l_subj_total = l_tj_fijado + delta_i_j

            self.tabla_sustitucion.item(
                item_sust_activo,
                values=(
                    f"{peso_bascula:.2f}",
                    f"{delta_i_j:.2f}",
                    f"{l_subj_total:.2f}",
                ),
            )

            if idx + 1 < len(filas_pat):
                self.tabla_patrones.selection_set(filas_pat[idx + 1])
                self.tabla_sustitucion.selection_set(filas_sust[idx + 1])
                self.tabla_patrones.see(filas_pat[idx + 1])
            return

        CTkMessagebox(
            title="Aviso",
            message="El Step seleccionado ya tiene todas sus lecturas completas.",
            icon="info",
        )

    def mostrar_selector_pesas(self):
        diccionario_pesas = getattr(self, "pesas", {})

        lista_patrones = [
            llave
            for llave, datos in diccionario_pesas.items()
            if str(datos.get("sustitucion")).strip().lower() == "no"
        ]

        if lista_patrones:
            Ventana_Pesas(self, lista_patrones, self.actualizar_tabla_pesas)
        else:
            CTkMessagebox(
                title="Sin registros",
                message="No se encontraron elementos marcados como patrones ('no') en la columna L del archivo cargado.",
                icon="warning",
            )

    def actualizar_tabla_pesas(self, seleccionadas):
        if not hasattr(self, "llaves_actuales") or self.llaves_actuales is None:
            self.llaves_actuales = []

        for llave in seleccionadas:
            if llave not in self.llaves_actuales:
                self.llaves_actuales.append(llave)

        for item in self.tabla_sel.get_children():
            self.tabla_sel.delete(item)

        suma_total = 0.0
        diccionario_pesas = getattr(self, "pesas", {})

        for i, llave in enumerate(self.llaves_actuales, start=1):
            datos = diccionario_pesas.get(llave)
            if datos:
                self.tabla_sel.insert(
                    "",
                    "end",
                    values=(
                        i,
                        datos["id"],
                        datos["serie"],
                        datos["nominal"],
                        datos["unidad"],
                    ),
                )

                try:
                    val = float(datos["nominal"])
                    unidad = str(datos["unidad"]).lower().strip()
                    if unidad in ["g", "gramos", "g."]:
                        suma_total += val / 1000.0
                    else:
                        suma_total += val
                except ValueError:
                    pass

        self.label_suma_nominal.configure(text=f"Total: {suma_total:.2f} kg")

    def mostrar_selector_sustitucion(self):
        diccionario_pesas = getattr(self, "pesas", {})

        lista_sustitucion = [
            llave
            for llave, datos in diccionario_pesas.items()
            if datos.get("sustitucion") == "si"
        ]

        if lista_sustitucion:
            Ventana_Pesas(self, lista_sustitucion, self.actualizar_tabla_pesas)
        else:
            CTkMessagebox(
                title="Sin registros",
                message="No se encontraron elementos marcados como sustitución en el archivo cargado.",
                icon="warning",
            )

    def borrar_punto(self):
        if not self.tabla.get_children():
            CTkMessagebox(
                title="Error", message="No hay puntos para eliminar.", icon="cancel"
            )
            return

        if (
            str(self.tabla.item(self.tabla.get_children()[-1])["values"][0]) != "0"
            and str(self.tabla.item(self.tabla.get_children()[-1])["values"][2]).strip()
            == "---"
            and str(self.tabla.item(self.tabla.get_children()[-1])["values"][4]).strip()
            == "---"
        ):
            self.tabla.delete(self.tabla.get_children()[-1])
            if len(self.tabla.get_children()) > 1:
                self.tabla.selection_set(self.tabla.get_children()[-2])
            return

    def eliminar_ultima_indicacion(self):
        filas = self.tabla.get_children()
        if not filas:
            return

        ultimo_item = filas[-1]
        v_ult = self.tabla.item(ultimo_item)["values"]
        if (
            str(v_ult[0]) != "0"
            and str(v_ult[2]).strip() == "---"
            and str(v_ult[4]).strip() == "---"
        ):
            self.tabla.delete(ultimo_item)
            if len(filas) > 1:
                self.tabla.selection_set(filas[-2])
            return

        for item in reversed(filas):
            v = self.tabla.item(item)["values"]
            if str(v[0]) == "0":
                continue

            if str(v[5]).strip() != "---":
                self.tabla.item(
                    item, values=(v[0], v[1], v[2], v[3], v[4], "---", "---", "---")
                )
                self.tabla.selection_set(item)
                return

            if str(v[4]).strip() != "---":
                self.tabla.item(
                    item, values=(v[0], v[1], v[2], v[3], "---", v[5], v[6], v[7])
                )
                self.tabla.selection_set(item)
                return

            if str(v[2]).strip() != "---":
                self.tabla.item(
                    item, values=(v[0], "---", "---", "---", "---", "---", "---", "---")
                )
                self.tabla.selection_set(item)
                return

        CTkMessagebox(
            title="Aviso", message="No hay indicaciones para eliminar.", icon="info"
        )

    def añadir_nuevo_step_dinamico(self):
        filas_pat = self.tabla_patrones.get_children()

        if not filas_pat:
            self.tabla_patrones.insert("", "end", values=("0", "0", "0", "0", "0"))
            self.tabla_sustitucion.insert("", "end", values=("0", "0", "0"))
            self.hora_inicio_prueba = datetime.now().strftime("%H:%M:%S")

            self.tabla_patrones.selection_set(self.tabla_patrones.get_children()[0])
            return

        ultimo_item_pat = filas_pat[-1]
        v_pat = self.tabla_patrones.item(ultimo_item_pat)["values"]

        filas_sust = self.tabla_sustitucion.get_children()
        ultimo_item_sust = filas_sust[-1]
        v_sust = self.tabla_sustitucion.item(ultimo_item_sust)["values"]

        if str(v_pat[0]) != "0":
            if str(v_pat[2]).strip() == "---" or str(v_pat[4]).strip() == "---":
                from CTkMessagebox import CTkMessagebox

                CTkMessagebox(
                    title="Paso Incompleto",
                    message=f"Debes registrar la carga (Iⱼ) y la descarga (I'ⱼ) del Step {v_pat[0]} antes de abrir uno nuevo.",
                    icon="warning",
                )
                return

        siguiente_j = int(v_pat[0]) + 1

        item_p = self.tabla_patrones.insert(
            "", "end", values=(str(siguiente_j), "---", "---", "---", "---")
        )
        item_s = self.tabla_sustitucion.insert("", "end", values=("---", "---", "---"))

        self.tabla_patrones.selection_set(item_p)
        self.tabla_patrones.see(item_p)

        self.tabla_sustitucion.selection_set(item_s)

    def registrar_sustitucion(self):
        self.sustitucion = not self.sustitucion

        if self.sustitucion:
            self.btn_registro_sustitucion.configure(
                text="Sustitución Manual Activa",
                fg_color="#1E8449",
                hover_color="#145A32",
            )
            CTkMessagebox(
                title="Sustitucion",
                message="Coloca el peso sobre la plataforma",
                icon="info",
            )

        else:
            self.btn_registro_sustitucion.configure(
                text="Sustitución Manual Desactivada",
                fg_color="#5D6D7E",
                hover_color="#34495E",
            )

    def borrar_ultima_indicacion(self):
        seleccion_actual = self.tabla_patrones.selection()
        if not seleccion_actual:
            from CTkMessagebox import CTkMessagebox

            CTkMessagebox(
                title="Aviso",
                message="Seleccione en la tabla para eliminar",
                icon="info",
            )
            return

        item_pat_activo = seleccion_actual[0]
        filas_pat = list(self.tabla_patrones.get_children())
        filas_sust = list(self.tabla_sustitucion.get_children())

        idx = filas_pat.index(item_pat_activo)
        item_sust_activo = filas_sust[idx]

        v_pat = self.tabla_patrones.item(item_pat_activo)["values"]
        v_sust = self.tabla_sustitucion.item(item_sust_activo)["values"]

        if str(v_pat[0]) == "0":
            return

        if str(v_sust[0]).strip() != "---":
            self.tabla_sustitucion.item(item_sust_activo, values=("---", "---", "---"))
            return

        if str(v_pat[4]).strip() != "---":
            self.tabla_patrones.item(
                item_pat_activo, values=(v_pat[0], v_pat[1], v_pat[2], v_pat[3], "---")
            )
            return

        if str(v_pat[2]).strip() != "---":
            self.tabla_patrones.item(
                item_pat_activo, values=(v_pat[0], "---", "---", "---", "---")
            )
            self.tabla_sustitucion.item(item_sust_activo, values=("---", "---", "---"))
            return

        from CTkMessagebox import CTkMessagebox

        CTkMessagebox(
            title="Información", message="Este Step ya se encuentra vacío.", icon="info"
        )

    def borrar_pesas(self):
        if not hasattr(self, "llaves_actuales") or not self.llaves_actuales:
            CTkMessagebox(title="Pesas", message="Tabla vacia", icon="info")
            return

        if self.tabla_sel.get_children():
            self.tabla_sel.delete(self.tabla_sel.get_children()[-1])

        self.llaves_actuales.pop()

        total = 0.0
        for k in self.llaves_actuales:
            if self.pesas.get(k):
                try:
                    v = float(self.pesas.get(k).get("nominal", 0))
                    valor = (
                        v / 1000.0
                        if str(self.pesas.get(k).get("unidad")).lower().strip()
                        in ["g", "g."]
                        else v
                    )
                    total += valor
                except ValueError:
                    pass

        self.label_suma_nominal.configure(text=f"Total: {total:.2f} kg")

    def registrar_CI(self):
        T = self.label_temperatura.cget("text")
        H = self.label_higrometro.cget("text")
        P = self.label_presion.cget("text")

        t_inst = self.combo_temperatura.get()
        h_inst = self.combo_higrometro.get()
        p_inst = self.combo_presion.get()

        self.condiciones_iniciales = {
            "Temperatura": {"Valor": T, "Instrumento": t_inst},
            "Humedad": {"Valor": H, "Instrumento": h_inst},
            "Presion": {"Valor": P, "Instrumento": p_inst},
            "Hora": datetime.now().strftime("%H:%M:%S"),
        }
        self.label_barI.configure(text=P, text_color="#158A30")
        self.label_higI.configure(text=H, text_color="#158A30")
        self.label_temI.configure(text=T, text_color="#158A30")

        self.btn_CI.configure(fg_color="#158A30")

    def registrar_CF(self):
        T = self.label_temperatura.cget("text")
        H = self.label_higrometro.cget("text")
        P = self.label_presion.cget("text")

        t_inst = self.combo_temperatura.get()
        h_inst = self.combo_higrometro.get()
        p_inst = self.combo_presion.get()

        self.condiciones_finales = {
            "Temperatura": {"Valor": T, "Instrumento": t_inst},
            "Humedad": {"Valor": H, "Instrumento": h_inst},
            "Presion": {"Valor": P, "Instrumento": p_inst},
            "Hora": datetime.now().strftime("%H:%M:%S"),
        }
        self.label_barF.configure(text=P, text_color="#D95E4A")
        self.label_higF.configure(text=H, text_color="#D95E4A")
        self.label_temF.configure(text=T, text_color="#D95E4A")

        self.btn_CF.configure(fg_color="#D95E4A")

    def guardar_xlsx(self):
        ruta_existente = getattr(self.master, "archivo_calibracion_actual", None)

        if ruta_existente and os.path.exists(ruta_existente):
            archivo = ruta_existente
            try:
                wb = openpyxl.load_workbook(archivo)
            except Exception as e:
                CTkMessagebox(
                    title="Error",
                    message=f"No se pudo abrir el archivo existente: {e}",
                    icon="cancel",
                )
                return
        else:
            hora = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            nombre_id = f"Ejercicio_Calibracion_{hora}.xlsx"

            archivo = asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=nombre_id,
                title="Guardar reporte",
            )
            if not archivo:
                return

            wb = openpyxl.Workbook()
            std = wb.get_sheet_by_name("Sheet") if "Sheet" in wb.sheetnames else None
            if std:
                wb.remove(std)

        try:
            nombre = "Error_Indicacion"
            contador = 1
            nuevo_nombre = nombre

            while nuevo_nombre in wb.sheetnames:
                contador += 1
                nuevo_nombre = f"{nombre} {contador}"

            hoja = wb.create_sheet(title=nuevo_nombre)

            color_encabezado = PatternFill(
                start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
            )
            fuente_negrita = Font(bold=True)

            def agregar_titulo(texto):
                hoja.append([texto])
                celda = hoja.cell(row=hoja.max_row, column=1)
                celda.fill = color_encabezado
                celda.font = fuente_negrita

            agregar_titulo("Indicacion")
            hoja.append(["#", "LTj", "Ij", "Ej", "I´j", "Isubj", "deltaIj", "Lsubj"])

            filas_pat = self.tabla_patrones.get_children()
            filas_sust = self.tabla_sustitucion.get_children()

            for idx, item_pat in enumerate(filas_pat):
                item_sust = filas_sust[idx]
                v_pat = self.tabla_patrones.item(item_pat)["values"]
                v_sust = self.tabla_sustitucion.item(item_sust)["values"]

                renglon_datos = [
                    v_pat[0],  # #
                    v_pat[1],  # LTj
                    v_pat[2],  # Ij
                    v_pat[3],  # Ej
                    v_pat[4],  # I´j
                    v_sust[0],  # Isubj
                    v_sust[1],  # deltaIj
                    v_sust[2],  # Lsubj
                ]

                renglon_procesado = []
                for val in renglon_datos:
                    try:
                        if str(val).strip() != "---":
                            renglon_procesado.append(
                                float(val) if "." in str(val) else int(val)
                            )
                        else:
                            renglon_procesado.append(val)
                    except ValueError:
                        renglon_procesado.append(val)

                hoja.append(renglon_procesado)
            hoja.append([])

            agregar_titulo("Pesas")
            hoja.append(
                [
                    "Key",
                    "Magnitud",
                    "Identificación",
                    "Modelo",
                    "Serie",
                    "Juego",
                    "Id_pesa",
                    "Nominal",
                    "Unidad",
                ]
            )

            if hasattr(self, "llaves_actuales"):
                for llave in self.llaves_actuales:
                    d = self.pesas.get(llave)
                    if d:
                        hoja.append(
                            [
                                d.get("key", "---"),
                                d.get("magnitud", "---"),
                                d.get("id", "---"),
                                d.get("modelo", "---"),
                                d.get("serie", "---"),
                                d.get("juego", "---"),
                                d.get("id_pesa", "---"),
                                d.get("nominal", 0),
                                d.get("unidad", "kg"),
                            ]
                        )
            hoja.append([])

            agregar_titulo("Medidores")
            combos = {
                "Presión": (
                    self.combo_presion.get()
                    if hasattr(self, "combo_presion")
                    else "---"
                ),
                "Humedad": (
                    self.combo_higrometro.get()
                    if hasattr(self, "combo_higrometro")
                    else "---"
                ),
                "Temperatura": (
                    self.combo_temperatura.get()
                    if hasattr(self, "combo_temperatura")
                    else "---"
                ),
            }
            for tipo, seleccion in combos.items():
                hoja.append([tipo, seleccion])
            hoja.append([])

            agregar_titulo("Condiciones_Ambientales")
            if hasattr(self, "condiciones_iniciales"):
                ci = self.condiciones_iniciales
                for mag in ["Presion", "Humedad", "Temperatura"]:
                    hoja.append(["CI", mag, ci[mag]["Valor"] if mag in ci else "---"])

            if hasattr(self, "condiciones_finales"):
                cf = self.condiciones_finales
                for mag in ["Presion", "Humedad", "Temperatura"]:
                    hoja.append(["CF", mag, cf[mag]["Valor"] if mag in cf else "---"])

            wb.save(archivo)

            if not ruta_existente:
                setattr(self.master, "archivo_calibracion_actual", archivo)

            CTkMessagebox(
                title="Guardar",
                message=f"Prueba añadida con éxito en la hoja '{nuevo_nombre}'.",
                icon="check",
            )

        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"No se pudo guardar el reporte: {e}",
                icon="cancel",
            )

    def abrir_xlsx(self):
        ruta = askopenfilename(
            title="Seleccionar archivo", filetypes=[("Excel files", "*.xlsx")]
        )
        if not ruta:
            return
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            hojas_error = [j for j in wb.sheetnames if "Error_Indicacion" in j]

            if not hojas_error:
                CTkMessagebox(
                    title="Error",
                    message="No se encontraron hojas de Error de Indicación.",
                    icon="cancel",
                )
                return

            if len(hojas_error) > 1:
                dialogo = ctk.CTkInputDialog(
                    text=f'Seleccion de hoja:\n ({" , ".join(hojas_error)})',
                    title="Seleccionar hoja",
                )
                nombre_hoja = dialogo.get_input()
                if nombre_hoja not in wb.sheetnames:
                    CTkMessagebox(
                        title="Error", message="Hoja no encontrada.", icon="cancel"
                    )
                    return
            else:
                nombre_hoja = hojas_error[0]

            hoja = wb[nombre_hoja]

            for item in self.tabla_patrones.get_children():
                self.tabla_patrones.delete(item)
            for item in self.tabla_sustitucion.get_children():
                self.tabla_sustitucion.delete(item)
            for item in self.tabla_sel.get_children():
                self.tabla_sel.delete(item)

            seccion = None
            llaves_pesas_a_cargar = []

            for fila in hoja.iter_rows(values_only=True):
                if not fila or all(c is None for c in fila):
                    seccion = None
                    continue

                primer_valor = str(fila[0]).strip()

                if primer_valor == "Indicacion":
                    seccion = "MEDICIONES"
                    continue
                elif primer_valor == "Pesas":
                    seccion = "PESAS"
                    continue
                elif primer_valor == "Medidores":
                    seccion = "MEDIDORES"
                    continue
                elif primer_valor == "Condiciones_Ambientales":
                    seccion = "CONDICIONES"
                    continue
                if seccion == "MEDICIONES" and primer_valor != "#":
                    try:
                        v = [str(c).strip() if c is not None else "---" for c in fila]

                        self.tabla_patrones.insert(
                            "", "end", values=(v[0], v[1], v[2], v[3], v[4])
                        )
                        self.tabla_sustitucion.insert(
                            "", "end", values=(v[5], v[6], v[7])
                        )
                    except:
                        continue

                elif seccion == "PESAS" and primer_valor != "Key":
                    llave_excel = primer_valor
                    if (
                        hasattr(self, "lista_pesas_completa")
                        and llave_excel in self.lista_pesas_completa
                    ):
                        llaves_pesas_a_cargar.append(llave_excel)
                    elif hasattr(self, "lista_pesas_completa"):
                        id_e = str(fila[2]).strip()
                        ser_e = str(fila[4]).strip()
                        for llave_m in self.lista_pesas_completa:
                            if id_e in llave_m and ser_e in llave_m:
                                llaves_pesas_a_cargar.append(llave_m)
                                break

                elif seccion == "MEDIDORES":
                    tipo, seleccion = primer_valor, str(fila[1])
                    if seleccion and seleccion != "None":
                        if "Presión" in tipo and hasattr(self, "combo_presion"):
                            self.combo_presion.set(seleccion)
                        elif "Humedad" in tipo and hasattr(self, "combo_higrometro"):
                            self.combo_higrometro.set(seleccion)
                        elif "Temperatura" in tipo and hasattr(
                            self, "combo_temperatura"
                        ):
                            self.combo_temperatura.set(seleccion)

                elif seccion == "CONDICIONES":
                    etapa, mag, valor = fila[0], fila[1], str(fila[2])
                    if etapa == "CI":
                        if "Presion" in mag and hasattr(self, "label_barI"):
                            self.label_barI.configure(text=valor, text_color="#158A30")
                        if "Humedad" in mag and hasattr(self, "label_higI"):
                            self.label_higI.configure(text=valor, text_color="#158A30")
                        if "Temperatura" in mag and hasattr(self, "label_temI"):
                            self.label_temI.configure(text=valor, text_color="#158A30")
                    elif etapa == "CF":
                        if "Presion" in mag and hasattr(self, "label_barF"):
                            self.label_barF.configure(text=valor, text_color="#D95E4A")
                        if "Humedad" in mag and hasattr(self, "label_higF"):
                            self.label_higF.configure(text=valor, text_color="#D95E4A")
                        if "Temperatura" in mag and hasattr(self, "label_temF"):
                            self.label_temF.configure(text=valor, text_color="#D95E4A")

            if llaves_pesas_a_cargar:
                self.actualizar_tabla_pesas(llaves_pesas_a_cargar)

            if hasattr(self, "desviacion"):
                self.desviacion()

            CTkMessagebox(
                title="Éxito", message="Reporte cargado correctamente", icon="check"
            )

        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"No se pudo cargar el archivo: {e}",
                icon="cancel",
            )

    def nueva_prueba(self):
        for i in self.tabla_sel.get_children():
            self.tabla_sel.delete(i)
        for j in self.tabla_sustitucion.get_children():
            self.tabla_sustitucion.delete(j)
        for k in self.tabla_patrones.get_children():
            self.tabla_patrones.delete(k)

        self.label_suma_nominal.configure()  # text="Suma Nominal: 0 kg")
        self.label_barI.configure()  # text="--", text_color="black")
        self.label_higI.configure()  # text="--", text_color="black")
        self.label_temI.configure()  # text="--", text_color="black")
        self.label_barF.configure()  # text="---", text_color="black")
        self.label_higF.configure()  # text="---", text_color="black")
        self.label_temF.configure()  # text="---", text_color="black")
        self.btn_CI.configure()  # )
        self.btn_CF.configure()

    def cargar_datos_desde_excel(self):
        ruta_directa = getattr(self.master, "archivo_calibracion_actual", None)
        if not ruta_directa:
            CTkMessagebox(
                title="Archivo Faltante",
                message="No hay ningún archivo cargado en el sistema.\n, ve al menú superior y carga primero.",
                icon="warning",
            )
            return

        try:
            wb = openpyxl.load_workbook(ruta_directa, data_only=True)
            hojas_error = [
                hoja for hoja in wb.sheetnames if hoja.startswith("Error_Indicacion")
            ]

            if not hojas_error:
                CTkMessagebox(
                    title="Hoja no encontrada",
                    message="El archivo cargado no contiene ninguna pestaña de 'Error de Indicación'.",
                    icon="info",
                )
                return

            hoja_seleccionada = hojas_error[0]
            if len(hojas_error) > 1:
                lista_hojas_str = "\n".join([f"- {h}" for h in hojas_error])
                dialogo = ctk.CTkInputDialog(
                    text=f"Se encontraron las siguientes pruebas:\n{lista_hojas_str}\n\nEscribe el nombre exacto de la prueba que deseas cargar:",
                    title="Seleccionar Prueba",
                )
                seleccion = dialogo.get_input()
                if seleccion and seleccion.strip() in wb.sheetnames:
                    hoja_seleccionada = seleccion.strip()
                elif seleccion is None or seleccion.strip() == "":
                    return
                else:
                    CTkMessagebox(
                        title="Error", message="Nombre inválido.", icon="cancel"
                    )
                    return

            hoja = wb[hoja_seleccionada]

            for i in self.tabla_patrones.get_children():
                self.tabla_patrones.delete(i)
            for i in self.tabla_sustitucion.get_children():
                self.tabla_sustitucion.delete(i)
            for i in self.tabla_sel.get_children():
                self.tabla_sel.delete(i)

            bloque_actual = None
            contador_pesas = 1
            suma_total_pesas = 0.0
            self.llaves_actuales = []

            for fila in hoja.iter_rows(values_only=True):
                if not fila or fila[0] is None:
                    continue
                primer_valor = str(fila[0]).strip()

                if primer_valor == "Indicacion":
                    bloque_actual = "INDICACION"
                    continue
                elif primer_valor == "Pesas":
                    bloque_actual = "PESAS"
                    continue
                elif primer_valor == "Medidores":
                    bloque_actual = "MEDIDORES"
                    continue
                elif primer_valor == "Condiciones_Ambientales":
                    bloque_actual = "CONDICIONES"
                    continue

                if bloque_actual == "INDICACION":
                    if primer_valor == "#":
                        continue

                    v = [str(c).strip() if c is not None else "---" for c in fila]

                    try:
                        self.tabla_patrones.insert(
                            "", "end", values=(v[0], v[1], v[2], v[3], v[4])
                        )
                        self.tabla_sustitucion.insert(
                            "", "end", values=(v[5], v[6], v[7])
                        )
                    except Exception:
                        continue

                elif bloque_actual == "PESAS":
                    if primer_valor in [
                        "Key",
                        "No se seleccionaron pesas para esta prueba",
                    ]:
                        continue

                    id_pesa = str(fila[2]).strip() if fila[2] is not None else ""
                    serie_pesa = str(fila[4]).strip() if fila[4] is not None else ""
                    nominal_pesa = fila[7] if fila[7] is not None else "0"
                    unidad_pesa = str(fila[8]).strip() if fila[8] is not None else "kg"
                    key_pesa = str(fila[0]).strip() if fila[0] is not None else ""

                    if key_pesa:
                        pesa_key_completa = (
                            f"{id_pesa} | {serie_pesa} | {nominal_pesa}{unidad_pesa}"
                        )
                        self.llaves_actuales.append(pesa_key_completa)

                    self.tabla_sel.insert(
                        "",
                        "end",
                        values=(
                            contador_pesas,
                            id_pesa,
                            serie_pesa,
                            nominal_pesa,
                            unidad_pesa,
                        ),
                    )
                    contador_pesas += 1

                    try:
                        val = float(nominal_pesa)
                        if unidad_pesa.lower() in ["g", "gramos", "g."]:
                            suma_total_pesas += val / 1000.0
                        else:
                            suma_total_pesas += val
                    except ValueError:
                        pass

                elif bloque_actual == "MEDIDORES":
                    tipo, seleccion = primer_valor, str(fila[1])
                    if seleccion and seleccion != "None":
                        if "Presión" in tipo and hasattr(self, "combo_presion"):
                            self.combo_presion.set(seleccion)
                        elif "Humedad" in tipo and hasattr(self, "combo_higrometro"):
                            self.combo_higrometro.set(seleccion)
                        elif "Temperatura" in tipo and hasattr(
                            self, "combo_temperatura"
                        ):
                            self.combo_temperatura.set(seleccion)

                elif bloque_actual == "CONDICIONES":
                    etapa, mag, valor = fila[0], fila[1], str(fila[2])
                    if etapa == "CI":
                        if "Presion" in mag and hasattr(self, "label_barI"):
                            self.label_barI.configure(text=valor, text_color="#158A30")
                        if "Humedad" in mag and hasattr(self, "label_higI"):
                            self.label_higI.configure(text=valor, text_color="#158A30")
                        if "Temperatura" in mag and hasattr(self, "label_temI"):
                            self.label_temI.configure(text=valor, text_color="#158A30")
                    elif etapa == "CF":
                        if "Presion" in mag and hasattr(self, "label_barF"):
                            self.label_barF.configure(text=valor, text_color="#D95E4A")
                        if "Humedad" in mag and hasattr(self, "label_higF"):
                            self.label_higF.configure(text=valor, text_color="#D95E4A")
                        if "Temperatura" in mag and hasattr(self, "label_temF"):
                            self.label_temF.configure(text=valor, text_color="#D95E4A")

            if hasattr(self, "label_suma_nominal"):
                self.label_suma_nominal.configure(
                    text=f"Total: {suma_total_pesas:.2f} kg"
                )

            if hasattr(self, "desviacion"):
                self.desviacion()

            CTkMessagebox(
                title="Éxito",
                message=f"Se cargaron correctamente los datos de '{hoja_seleccionada}'.",
                icon="check",
            )

        except Exception as e:
            CTkMessagebox(
                title="Error",
                message=f"No se pudieron cargar los datos de error de indicación:\n{str(e)}",
                icon="cancel",
            )


class Window(ctk.CTk):
    """
    clase principal de la ventana, aqui se crean los widgepesots, los menus y se actualiza el peso al momento
    """

    def __init__(self, comunicacion: Comunication, shell: Shell):
        super().__init__()

        self.shell = shell
        self.ultimo_peso = 0
        self.comunicacion = comunicacion
        self.tara_activa = False
        self.val_tara = 0

        # ---------------------------
        self.peso_maximo = 58000
        self.alerta_peso_maximo = False
        # ---------------------------

        self.title("MT")
        self.geometry("1000x600+100+100")

        if os.path.exists(
            ruta(
                os.path.join(
                    "Icon",
                    "CNM_color_100dpi.ico",
                )
            )
            # r"Icon\weight_fat_body_overweight_unhealthy_obesity_diet_health_belly_icon_133507.ico"
        ):
            self.iconbitmap(
                ruta(
                    os.path.join(
                        "Icon",
                        "CNM_color_100dpi.ico",
                    )
                )
                # r"Icon\weight_fat_body_overweight_unhealthy_obesity_diet_health_belly_icon_133507.ico"
            )
        self.b_menus()
        self.widgets()
        # self.b_menus()
        # self.barra_menus()
        self.peso_al_momento()

    def widgets(self):
        """
        Son los metodos propios de la ventana
        """
        self.tabla_frame = ctk.CTkFrame(self, width=400, height=600)
        self.tabla_frame.pack(side="left", fill="both", padx=10, pady=10)

        # ------------TABLAAA----------------------
        self.style_tabla = ttk.Style()
        self.style_tabla.theme_use("clam")
        self.style_tabla.configure(
            "Treeview",
            background="#BD9D9D",
            foreground="black",
            rowheight=25,
            fieldbackground="#DFF2F5",
            font=("Cambria", 10),
        )

        self.tabla = ttk.Treeview(
            self.tabla_frame, columns=("incremento", "indicacion", "fecha", "hora")
        )

        self.tabla.heading("incremento", text="Incremento")
        self.tabla.heading("indicacion", text="Indicación")
        self.tabla.heading("fecha", text="Fecha")
        self.tabla.heading("hora", text="Hora")
        self.tabla.heading("#0", text="#")

        self.tabla.column("incremento", width=80, anchor="center")
        self.tabla.column("indicacion", width=100, anchor="center")
        self.tabla.column("fecha", width=100, anchor="center")
        self.tabla.column("hora", width=100, anchor="center")
        self.tabla.column("#0", width=40, anchor="center")

        self.tabla.pack(padx=20, pady=20)

        # ---------------BOTON DE REGISTRO--------------
        imagen_registro = Image.open(ruta(os.path.join("Icon", "boton-agregar.png")))
        # imagen_registro = Image.open(r"Icon\boton-agregar.png")
        icon_registro = ctk.CTkImage(
            light_image=imagen_registro, dark_image=imagen_registro, size=(30, 30)
        )

        self.button_registro = ctk.CTkButton(
            self.tabla_frame,
            text="Registrar",
            width=3,
            height=20,
            corner_radius=20,
            border_color="white",
            command=self.registrar_peso,
            image=icon_registro,
            compound="top",
        )
        self.button_registro.pack()

        # ----------- Frame Bottones de la Tabla-----------

        self.frame_buttons_tabla = ctk.CTkFrame(self.tabla_frame, width=400, height=150)
        self.frame_buttons_tabla.pack(expand=True)
        self.frame_buttons_tabla.pack_propagate(False)

        # ------------Botones de la tabla-------------------
        """
        Botones de la Tabla nuevo, eliminar etc
        """
        # --------ARREGLAR LOS BOTONES--------------------

        # +++++++++++++Boton guardar++++++++++++++
        # imagen_guardar = Image.open(r"Icon\guardar-el-archivo.png")
        imagen_guardar = Image.open(
            ruta(os.path.join("Icon", "guardar-el-archivo.png"))
        )
        icon_save = ctk.CTkImage(
            light_image=imagen_guardar, dark_image=imagen_guardar, size=(30, 30)
        )

        self.button_guardar = ctk.CTkButton(
            self.frame_buttons_tabla,
            text="Guardar",
            width=3,
            corner_radius=20,
            border_width=1,
            border_color="white",
            command=self.guardar_xlsx,
            image=icon_save,
            compound="top",
        )
        self.button_guardar.pack(side="left", expand=True)
        # +++++++++++++++++++++++++++++++++++++++++

        # -----------BOTON NUEVO-------------------
        imagen_new = Image.open(ruta(os.path.join("Icon", "anadir.png")))
        # imagen_new = Image.open(r"Icon\anadir.png")
        icon_new = ctk.CTkImage(
            light_image=imagen_new, dark_image=imagen_new, size=(30, 30)
        )
        self.button_nuevo = ctk.CTkButton(
            self.frame_buttons_tabla,
            text="Nuevo",
            width=3,
            height=20,
            corner_radius=20,
            border_width=1,
            border_color="white",
            command=self.limpiar,
            image=icon_new,
            compound="top",
        )
        self.button_nuevo.pack(side="left", expand=True)
        # -------------------------------------------------
        imagen_eliminar_ultimo = Image.open(ruta(os.path.join("Icon", "borrar.png")))
        # imagen_eliminar_ultimo = Image.open(r"Icon/borrar.png")
        icono_eliminar_ultimo = ctk.CTkImage(
            light_image=imagen_eliminar_ultimo,
            dark_image=imagen_eliminar_ultimo,
            size=(30, 30),
        )

        self.button_eliminar_ultimo = ctk.CTkButton(
            self.frame_buttons_tabla,
            text="Eliminar Ultimo",
            width=3,
            height=20,
            corner_radius=20,
            border_width=1,
            border_color="white",
            command=self.eliminar_ultimo,
            image=icono_eliminar_ultimo,
            compound="top",
        )
        self.button_eliminar_ultimo.pack(side="left", expand=True)
        # ******************************************************************

        # ------------------Frame PARA EL PUERTO-------------------------------

        # -----------------------Frame Principal-----------------------

        self.principal_frame = ctk.CTkFrame(self, width=50, height=60)
        self.principal_frame.pack(padx=10, pady=10, fill="both", expand=True)

        self.frame_peso = ctk.CTkFrame(self.principal_frame, width=300, height=120)
        self.frame_peso.pack(pady=40, padx=20, fill="both")
        self.frame_peso.pack_propagate(False)

        self.label_unidad = ctk.CTkLabel(
            self.frame_peso, text="kg", font=("Cambria", 50)
        )
        self.label_unidad.pack(side="right", padx=30)

        # ***********************************+

        self.frame_peso_al_momento = ctk.CTkFrame(
            self.frame_peso, width=400, height=150
        )
        self.frame_peso_al_momento.pack(pady=10, padx=10, fill="both", expand=True)
        self.frame_peso_al_momento.pack_propagate(False)

        self.frame_peso_al_momento.grid_columnconfigure(0, weight=1)
        self.frame_peso_al_momento.grid_columnconfigure(1, weight=3)
        self.frame_peso_al_momento.grid_rowconfigure((0, 1), weight=1)

        self.label_tara = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 15)
        )
        self.label_tara.grid(row=0, column=0, sticky="nw", padx=10, pady=10)

        self.label_neto = ctk.CTkLabel(
            self.frame_peso_al_momento, text="", font=("Cambria", 15)
        )
        self.label_neto.grid(row=1, column=0, sticky="sw", padx=10, pady=5)

        self.label_peso = ctk.CTkLabel(
            self.frame_peso_al_momento,
            text="",
            text_color="red",
            font=("Cambria", 90),
        )
        self.label_peso.grid(row=0, column=1, rowspan=2, sticky="nsew")

        # ----------- Frame de Botones Principales, tarar, zero etc

        self.frame_buttons_principales = ctk.CTkFrame(
            self.principal_frame, width=400, height=100
        )
        self.frame_buttons_principales.pack(pady=5)
        self.frame_buttons_principales.pack_propagate(False)

        # -----------BOTON DE TARAR--------------------
        imagen_tarar = Image.open(ruta(os.path.join("Icon", "aumentar.png")))
        # imagen_tarar = Image.open(r"Icon\aumentar.png")
        icon_tarar = ctk.CTkImage(
            light_image=imagen_tarar, dark_image=imagen_tarar, size=(30, 30)
        )

        self.button_tarar = ctk.CTkButton(
            self.frame_buttons_principales,
            text="Tara",
            width=3,
            height=20,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.tara,
            image=icon_tarar,
            compound="top",
        )
        self.button_tarar.pack(side="left", expand=True)

        # ------------BOTON DE ZERO-------------
        imagen_zero = Image.open(ruta(os.path.join("Icon", "cero.png")))
        # imagen_zero = Image.open(r"Icon\cero.png")
        icon_zero = ctk.CTkImage(
            light_image=imagen_zero, dark_image=imagen_zero, size=(30, 30)
        )

        self.button_zero = ctk.CTkButton(
            self.frame_buttons_principales,
            text="Zero",
            width=3,
            height=20,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.zero,
            image=icon_zero,
            compound="top",
        )
        self.button_zero.pack(side="left", expand=True)
        # -----------BOTON DE QUITAR TARA---------------

        imagen_q_tara = Image.open(ruta(os.path.join("Icon", "perdida-peso.png")))
        # imagen_q_tara = Image.open(r"Icon\perdida-peso.png")
        icon_q_tara = ctk.CTkImage(
            light_image=imagen_q_tara, dark_image=imagen_q_tara, size=(30, 30)
        )
        self.button_quitar_tara = ctk.CTkButton(
            self.frame_buttons_principales,
            text="Eliminar Tara",
            width=3,
            height=20,
            corner_radius=40,
            border_color="white",
            border_width=1,
            command=self.shell.quitar_tara,
            image=icon_q_tara,
            compound="top",
        )
        self.button_quitar_tara.pack(side="left", expand=True)

        # ------------FRAME PARA MOSTRAR EL PUERTO----------------

        self.frame_puerto = ctk.CTkFrame(self.principal_frame, width=300, height=50)
        self.frame_puerto.pack(side="bottom", pady=10)
        self.frame_puerto.pack_propagate(False)

        self.label_port = ctk.CTkLabel(
            self.frame_puerto, text="Puerto", font=("Cambria", 15)
        )
        self.label_port.pack(expand=True)

    def b_menus(self):
        """
        Crear los menus genericos
        """
        self.menu = CTkMenuBar(master=self)
        self.b1 = self.menu.add_cascade("Archivo")  # Menu de archivo
        self.b2 = self.menu.add_cascade("Puerto")  # Menu de puertos
        self.b3 = self.menu.add_cascade("Excentricidad", postcommand=self.Excentricidad)
        self.b4 = self.menu.add_cascade(
            "Repetibilidad", command=self.ventana_repetibilidad
        )
        self.b5 = self.menu.add_cascade(
            "Error de Indicación", command=self.Error_Indicacion
        )

        # -------------Las opciones solo del menu archivo---------------
        self.dropdown_1 = CustomDropdownMenu(widget=self.b1)
        self.dropdown_1.add_option(
            option="Abrir",
            command=self.abrir_xlsx,
            icon=ruta(os.path.join("Icon", "carpeta-abierta.png")),
            # icon=r"Icon\carpeta-abierta.png",
            # icon=r"Icon\folder_open_24dp_E3E3E3_FILL0_wght400_GRAD0_opsz24.png",
        )
        self.dropdown_1.add_option(
            option="Guardar",
            command=self.guardar_xlsx,
            icon=ruta(os.path.join("Icon", "marcador.png")),
            # icon=r"Icon\marcador.png",
        )
        self.dropdown_1.add_option(
            option="Limpiar tabla",
            command=self.limpiar,
            icon=ruta(os.path.join("Icon", "cepillar.png")),
            # icon=r"Icon\cepillar.png",
        )
        self.dropdown_1.add_option(
            option="Salir",
            command=self.destroy,
            icon=ruta(os.path.join("Icon", "salida.png")),
        )
        self.actualizar_puerto()
        self.dropdown_puerto.add_option(
            option="Cerrar puerto",
            command=self.comunicacion.cerrar_puerto,
            icon=ruta(os.path.join("Icon", "cable-roto.png")),
            # icon=r"Icon\desenchufado.png",
        )

        # self.dropdown_3 = CustomDropdownMenu(widget=self.b3)
        # self.dropdown_3.add_option(
        #     option="Ajustes",
        #     command=self.Excentricidad,
        #     icon=ruta(os.path.join("Icon", "negocio.png")),
        # )

    def actualizar_puerto(self):
        self.dropdown_puerto = CustomDropdownMenu(widget=self.b2)
        puertos = serial.tools.list_ports.comports()
        if not puertos:
            self.dropdown_puerto.add_option(option="Sin puertos")
        else:
            for p in puertos:
                self.dropdown_puerto.add_option(
                    option=p.device,
                    command=lambda puerto=p.device: self.seleccionar_puerto(puerto),
                    icon=ruta(os.path.join("Icon", "vga_cable.png")),
                    # icon=r"Icon\vga_cable.png",
                )

    def seleccionar_puerto(self, puerto_elegido):
        print(f"puerto elegido {puerto_elegido}")
        self.comunicacion.port = puerto_elegido

        self.label_port.configure(
            text=f"Puerto: {self.comunicacion.port} | {self.comunicacion.baud}"
        )
        if self.comunicacion.conexion():
            # inf = f"conexion: {self.comunicacion.port} | baud: {self.comunicacion.baud}"
            self.label_port.configure(text_color="#4E915D")
            print("conectado")
        else:
            self.label_port.configure(text_color="#823131")
            print("port fallo")

    def conectar_puerto(self, info):
        """
        Conexion del puerto al seleccionar
        """
        try:
            if hasattr(self, "comunicacion") and self.comunicacion:
                self.comunicacion.cerrar_puerto()

            self.comunicacion = Comunication(port=info["Dispositivo"], baud=9600)
            self.comunicacion.conexion()
            if self.comunicacion.conectado:
                self.shell = Shell(self.comunicacion)

                CTkMessagebox(title="Conexion", message="Conectado", icon="check")
            else:
                CTkMessagebox(title="Conexion", message="Salio mal", icon="warning")

        except Exception as e:
            CTkMessagebox(
                title="Error conexion", message="Error al conectar", icon="warning"
            )

    def registrar_peso(self):
        """
        Registra el peso con el button siempre que la lectura sea estable
        """
        lecture = self.shell.peso_estable()
        if len(lecture) >= 3:
            peso_actual = float(lecture[2])
            incremento = peso_actual - self.ultimo_peso

            fecha = datetime.now().strftime("%d/%m/%Y")
            hora = datetime.now().strftime("%H:%M:%S")
            id_fila = len(self.tabla.get_children()) + 1
            self.tabla.insert(
                "",
                "end",
                text=str(id_fila),
                values=(incremento, peso_actual, fecha, hora),
            )
            # self.table.insert('', 'end', values=(id_incremento,incremento, fecha, hora))
            self.ultimo_peso = peso_actual
            self.tabla.yview_moveto(1)

    def guardar_xlsx(self):
        """
        Guarda los registros de la tabla en archivo excel
        """

        if not self.tabla.get_children():
            CTkMessagebox(title="Erro", message="Sin datos", icon="warning")
            return

        archivo = asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("Archivo csv", "*.csv")],
            title="Pesossss",
        )
        if archivo:
            try:
                with open(archivo, mode="w", newline="", encoding="utf-8") as f:
                    escritor = csv.writer(f)
                    escritor.writerow(["Incremento", "Indicación", "Fecha", "Hora"])

                    for i in self.tabla.get_children():
                        val = self.tabla.item(i)["values"]
                        escritor.writerow(val)
                CTkMessagebox(title="Guardar", message="Guardado", icon="check")
            except Exception as e:
                print("Error")

    def abrir_xlsx(self):
        """
        Abre archivo con extension excel dentro de la interfaz
        """
        # if not self.table.get_children():
        #     CTkMessagebox(title='Erro',message='Sin datos', icon='warning' )
        #     return

        archivo = askopenfilename(
            filetypes=[("Archivo csv", "*.csv")], title="Pesossss"
        )
        if archivo:
            try:
                # self.limpiar()
                with open(archivo, mode="r", encoding="utf-8") as f:
                    lecture = csv.DictReader(f)
                    # escritor.writerow(['Incremento', 'Peso', 'Fecha','Hora'])

                    for i in lecture:
                        inc = i["Incremento"]
                        peso = i["Indicación"]
                        fecha = i["Fecha"]
                        hora = i["Hora"]

                        id_i = len(self.tabla.get_children()) + 1
                        self.tabla.insert(
                            "", "end", text=str(id_i), values=(inc, peso, fecha, hora)
                        )

                        items = self.tabla.get_children()
                        if items:
                            last_i = items[-1]
                            val = self.tabla.item(last_i)["values"]
                            self.ultimo_peso = float(val[1])
                    CTkMessagebox(title="Open", message="Exito", icon="check")
                    self.tabla.yview_moveto(1)
            except Exception as e:
                CTkMessagebox(title="Open", message="Warning", icon="warning")

    def eliminar_ultimo(self):
        """
        Elimina el ultimo elemento de la tabla
        """
        it = self.tabla.get_children()
        if not self.tabla.get_children():
            CTkMessagebox(title="Eliminar ultimo", message="Sin elementos")
            return
        i_last = it[-1]
        self.tabla.delete(i_last)

        new_i = self.tabla.get_children()

        if new_i:
            new_last = new_i[-1]
            val = self.tabla.item(new_last)["values"]
            self.ultimo_peso = float(val[1])
        else:
            self.ultimo_peso = 0

        CTkMessagebox(title="Eliminar ultimo", message="Eliminado", icon="check")

    def limpiar(self):
        """
        Elimina/limpia toda la tabla con los elementos
        """
        # self.tabla.delete(*self.tabla.get_children()) + 1
        for i in self.tabla.get_children():
            self.tabla.delete(i)

        self.ultimo_peso = 0

    def peso_al_momento(self):
        """
        Es la actualizacion del peso en tiempo real que se ve en la interfaz
        """
        lecture = self.shell.peso_instantaneo()

        tara_lecture = self.shell.obtener_tara()

        if len(lecture) >= 3:
            try:
                peso_r = int(float(lecture[2]))
                self.label_peso.configure(text=str(peso_r))
                self.label_neto.configure(
                    text=f"Neto: {peso_r} kg", text_color="#FF6EDE"
                )
            except ValueError:
                pass

        if "+" in lecture:
            self.label_peso.configure(text="----", text_color="red")

            if not self.alerta_peso_maximo:
                self.alerta_peso_maximo = True
                CTkMessagebox(
                    title="Maximo peso",
                    message="Sobrecarga. Retire peso",
                    icon="cancel",
                    option_1="aceptar",
                )
        elif len(lecture) >= 3:
            self.alerta_peso_maximo = False

        if len(lecture) >= 3:
            wg = lecture[2]
            self.label_peso.configure(text=wg)

            if lecture[1] == "S":
                self.label_peso.configure(text_color="#158A30")
            else:
                self.label_peso.configure(text_color="#B52C19")

            valor_tara = self.shell.obtener_tara()

            self.label_tara.configure(text=f"Tara: {valor_tara}", text_color="#FF6021")
        self.after(100, self.peso_al_momento)

    def ventana_repetibilidad(self):
        equipos = getattr(self, "lista_medidores", ["--"])

        if not hasattr(self, "v_repet") or not self.v_repet.winfo_exists():
            self.v_repet = Repetibilidad(self, self.shell, lista_instrumentos=equipos)
        else:
            self.v_repet.focus()

    def cargar_excel_calibracion(self):
        archivo = askopenfilename(
            title="Seleccionar Excel de Calibración",
            filetypes=[("Archivos de Excel", "*.xlsx")],
        )
        if archivo:
            self.archivo_calibracion_actual = archivo
            CTkMessagebox(
                title="Archivo Cargado",
                message=f"Archivo cargado con éxito:\n{os.path.basename(archivo)}",
                icon="check",
            )

            if (
                hasattr(self, "calibracion_window")
                and self.calibracion_window.winfo_exists()
            ):
                self.calibracion_window.cargar_datos_desde_excel()

    def Excentricidad(self):
        equipos = getattr(self, "lista_medidores", ["--"])

        ruta_excel = getattr(self, "archivo_calibracion_actual", None)

        if not hasattr(self, "v_excen") or not self.v_excen.winfo_exists():
            self.v_excen = Excentricidad(
                self,
                self.shell,
                lista_instrumentos=equipos,
                ruta_excel_trabajo=ruta_excel,
            )
        else:
            self.v_excen.focus()

    def Error_Indicacion(self):
        equipos = getattr(self, "lista_medidores", ["--"])

        ruta_excel = getattr(self, "archivo_calibracion_actual", None)

        if not hasattr(self, "ventana_error") or not self.ventana_error.winfo_exists():
            self.ventana_error = Error_Indicacion(
                parent=self,
                master=self,
                shell=self.shell,
                lista_instrumentos=equipos,
                ruta_excel=ruta_excel,
            )
        else:
            self.ventana_error.focus()
